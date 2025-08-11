"""
AREA DE CARTERA - PROCESO FORMATO DEUDA

Este script procesa el archivo de provisión (provision.csv) generado por el sistema Pisa, siguiendo las reglas y transformaciones requeridas por el área de cartera:

OBJETIVO:
El objetivo de implementar el formato de deuda es contar con la información necesaria de manera óptima, veraz y oportuna, lo que permitirá elaborar de forma eficiente el reporte requerido por la casa matriz.

PROCESO:
1. Renombra los campos según el mapeo oficial
2. Elimina la columna PCIMCO
3. Elimina la fila de empresa PL30 con valor -614.000
4. Unifica nombres de clientes en una sola columna
5. Convierte fechas y crea columnas de día, mes, año
6. Calcula días vencidos, días por vencer, saldo vencido
7. Calcula % dotación, valor dotación, mora total
8. Crea vencimientos históricos de los últimos 6 meses
9. Calcula vencimiento 180 días
10. Calcula valores por vencer de los próximos 3 meses
11. Calcula mayor a 90 días y valor total por vencer
12. Valida que la suma de mora total + valor por vencer sea igual al saldo
13. Crea columnas de vencimientos por rango de días y valida su suma
14. Crea columna de deuda incobrable
15. Aplica formato colombiano a los números

Este script procesa únicamente el archivo de provisión de forma independiente.
"""
import pandas as pd
import numpy as np
from datetime import datetime, date
import os
import sys
import locale
import warnings
warnings.filterwarnings('ignore')

# Mapeo oficial de columnas según especificaciones
MAPEO_PROVISION = {
    'PCCDEM': 'EMPRESA',
    'PCCDAC': 'ACTIVIDAD', 
    'PCDEAC': 'EMPRESA',
    'PCCDAG': 'CODIGO AGENTE',
    'PCNMAG': 'AGENTE',
    'PCCDCO': 'CODIGO COBRADOR',
    'PCNMCO': 'COBRADOR',
    'PCCDCL': 'CODIGO CLIENTE',
    'PCCDDN': 'IDENTIFICACION',
    'PCNMCL': 'NOMBRE',
    'PCNMCM': 'DENOMINACION COMERCIAL',
    'PCNMDO': 'DIRECCION',
    'PCTLF1': 'TELEFONO',
    'PCNMPO': 'CIUDAD',
    'PCNUFC': 'NUMERO FACTURA',
    'PCORPD': 'TIPO',
    'PCFEFA': 'FECHA',
    'PCFEVE': 'FECHA VTO',
    'PCVAFA': 'VALOR',
    'PCSALD': 'SALDO'
}

# Configuración de rangos de vencimiento según especificaciones
VENCIMIENTOS_RANGOS = [
    ('SALDO NO VENCIDO', 0, 29),
    ('VENCIDO 30', 30, 59),
    ('VENCIDO 60', 60, 89),
    ('VENCIDO 90', 90, 179),
    ('VENCIDO 180', 180, 359),
    ('VENCIDO 360', 360, 369),
    ('VENCIDO + 360', 370, 99999)
]

def obtener_fecha_cierre(fecha_cierre_str=None):
    """Obtiene la fecha de cierre. Si se proporciona fecha_cierre_str, la usa; si no, usa el último día del mes actual"""
    if fecha_cierre_str:
        try:
            return datetime.strptime(fecha_cierre_str, '%Y-%m-%d')
        except ValueError:
            print(f"ADVERTENCIA: Formato de fecha incorrecto '{fecha_cierre_str}'. Usando fecha por defecto.")
    
    # Fecha por defecto: último día del mes actual
    hoy = datetime.now()
    if hoy.month == 12:
        cierre = datetime(hoy.year + 1, 1, 1) - pd.Timedelta(days=1)
    else:
        cierre = datetime(hoy.year, hoy.month + 1, 1) - pd.Timedelta(days=1)
    return cierre

def limpiar_y_validar_datos(df):
    """Limpia y valida los datos del DataFrame"""
    print("Iniciando limpieza y validación de datos...")
    
    # Limpiar nombres de columnas
    df.columns = df.columns.str.strip()
    
    # Corregir nombres de columnas con caracteres especiales
    if 'DENOMINACIÓN COMERCIAL' in df.columns:
        df.rename(columns={'DENOMINACIÓN COMERCIAL': 'DENOMINACION COMERCIAL'}, inplace=True)
    
    # Renombrar columnas según mapeo oficial
    columnas_renombradas = {}
    for col_original, col_nueva in MAPEO_PROVISION.items():
        if col_original in df.columns:
            columnas_renombradas[col_original] = col_nueva
    
    df = df.rename(columns=columnas_renombradas)
    print(f"Columnas renombradas: {len(columnas_renombradas)}")
    
    # Eliminar columna PCIMCO si existe
    if 'PCIMCO' in df.columns:
        df = df.drop(columns=['PCIMCO'])
        print("Columna PCIMCO eliminada")
    
    # Eliminar fila de empresa PL30 (PCCDAC = 30 y valor -614.000)
    if 'ACTIVIDAD' in df.columns and 'SALDO' in df.columns:
        registros_antes = len(df)
        # Convertir valores de saldo para comparación
        saldos_convertidos = df['SALDO'].apply(lambda x: float(str(x).replace(',','').replace('$','')) if pd.notna(x) and str(x).replace(',','').replace('$','').replace('.','',1).replace('-','',1).isdigit() else 0)
        df = df[~((df['ACTIVIDAD'].astype(str).str.strip() == '30') & (saldos_convertidos == -614000))]
        registros_eliminados = registros_antes - len(df)
        if registros_eliminados > 0:
            print(f"Eliminados {registros_eliminados} registros de empresa PL30")
    
    # Validar y corregir valores negativos en saldos
    if 'SALDO' in df.columns:
        saldos_convertidos = df['SALDO'].apply(lambda x: float(str(x).replace(',','').replace('$','')) if pd.notna(x) and str(x).replace(',','').replace('$','').replace('.','',1).replace('-','',1).isdigit() else 0)
        valores_negativos = saldos_convertidos < 0
        if valores_negativos.any():
            print(f"ADVERTENCIA: Se encontraron {valores_negativos.sum()} registros con valores negativos en SALDO")
            print("Los valores negativos se convertirán a positivos para el procesamiento")
            df['SALDO'] = df['SALDO'].apply(lambda x: str(abs(float(str(x).replace(',','').replace('$','')))) if pd.notna(x) and float(str(x).replace(',','').replace('$','')) < 0 else str(x))
    
    return df

def unificar_nombres_clientes(df):
    """Unifica los nombres de clientes en una sola columna"""
    print("Unificando nombres de clientes...")
    
    if 'NOMBRE' in df.columns and 'DENOMINACION COMERCIAL' in df.columns:
        # Llenar valores vacíos en DENOMINACION COMERCIAL con NOMBRE
        df['DENOMINACION COMERCIAL'] = df['DENOMINACION COMERCIAL'].fillna('')
        df['NOMBRE'] = df['NOMBRE'].fillna('')
        
        # Unificar: si DENOMINACION COMERCIAL está vacía, usar NOMBRE
        df['DENOMINACION COMERCIAL'] = df.apply(
            lambda row: row['NOMBRE'] if pd.isna(row['DENOMINACION COMERCIAL']) or str(row['DENOMINACION COMERCIAL']).strip() == '' 
            else row['DENOMINACION COMERCIAL'], axis=1
        )
        
        print("Nombres de clientes unificados correctamente")
    
    return df

def procesar_fechas(df, fecha_cierre_str=None):
    """Procesa las fechas y crea columnas separadas"""
    print("Procesando fechas...")
    
    fecha_cierre = obtener_fecha_cierre(fecha_cierre_str)
    
    def convertir_fecha(fecha_str):
        """
        Convierte una cadena de fecha a formato (True, dia, mes, año, datetime) o (False, None, None, None, None) si falla.
        """
        if pd.isna(fecha_str) or str(fecha_str).strip() == '':
            return (False, None, None, None, None)
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
            try:
                fecha_dt = datetime.strptime(str(fecha_str).strip(), fmt)
                return (True, fecha_dt.day, fecha_dt.month, fecha_dt.year, fecha_dt)
            except Exception:
                continue
        return (False, None, None, None, None)

    for col_fecha in ['FECHA', 'FECHA VTO']:
        if col_fecha in df.columns:
            # Convertir fechas
            fechas_convertidas = df[col_fecha].apply(convertir_fecha)
            
            # Formatear fechas en formato dd/mm/yyyy
            fechas_formateadas = []
            dias = []
            meses = []
            años = []
            fechas_datetime = []
            
            for fecha_conv in fechas_convertidas:
                if fecha_conv[0]:  # Si la conversión fue exitosa
                    _, dia, mes, anio, fecha_dt = fecha_conv
                    fechas_formateadas.append(f"{dia:02d}/{mes:02d}/{anio}")
                    dias.append(dia)
                    meses.append(mes)
                    años.append(anio)
                    fechas_datetime.append(fecha_dt)
                else:
                    fechas_formateadas.append("")
                    dias.append("")
                    meses.append("")
                    años.append("")
                    fechas_datetime.append(None)
            
            # Actualizar columna original
            df[col_fecha] = fechas_formateadas
            
            # Crear columnas separadas
            df[f'DIA {col_fecha}'] = dias
            df[f'MES {col_fecha}'] = meses
            df[f'AÑO {col_fecha}'] = años
            
            # Guardar fechas como datetime para cálculos
            df[f'{col_fecha}_DT'] = fechas_datetime
    
    print("Fechas procesadas correctamente")
    return df

def calcular_dias_vencidos(df, fecha_cierre_str=None):
    """Calcula días vencidos y días por vencer"""
    print("Calculando días vencidos...")
    
    fecha_cierre = obtener_fecha_cierre(fecha_cierre_str)
    
    if 'FECHA VTO_DT' in df.columns and 'SALDO' in df.columns:
        dias_vencidos = []
        dias_por_vencer = []
        
        for fecha_vto in df['FECHA VTO_DT']:
            if fecha_vto and pd.notna(fecha_vto):
                dias_diff = (fecha_vto - fecha_cierre).days
                if dias_diff < 0:
                    # Vencido
                    dias_vencidos.append(abs(dias_diff))
                    dias_por_vencer.append(0)
                else:
                    # Por vencer
                    dias_vencidos.append(0)
                    dias_por_vencer.append(dias_diff)
            else:
                dias_vencidos.append(0)
                dias_por_vencer.append(0)
        
        df['DIAS VENCIDO'] = dias_vencidos
        df['DIAS POR VENCER'] = dias_por_vencer
        
        print("Días vencidos y por vencer calculados correctamente")
    
    return df

def calcular_saldos_y_dotacion(df):
    """Calcula saldo vencido, dotación y mora total"""
    print("Calculando saldos y dotación...")
    
    if 'SALDO' in df.columns and 'DIAS VENCIDO' in df.columns:
        # Saldo vencido
        df['SALDO VENCIDO'] = df.apply(
            lambda row: float(str(row['SALDO']).replace(',','').replace('$','')) if row['DIAS VENCIDO'] > 0 else 0, axis=1
        )
        
        # % Dotación (100% si días vencidos >= 180)
        df['% Dotación'] = df['DIAS VENCIDO'].apply(lambda x: '100%' if x >= 180 else '0%')
        
        # Valor Dotación (saldo si días vencidos >= 180)
        df['  Valor Dotación  '] = df.apply(
            lambda row: float(str(row['SALDO']).replace(',','').replace('$','')) if row['DIAS VENCIDO'] >= 180 else 0, axis=1
        )
        
        # Mora Total (igual al saldo vencido)
        df['Mora Total'] = df['SALDO VENCIDO']
        
        # Valor Total Por Vencer
        df['Valor Total Por Vencer'] = df.apply(
            lambda row: float(str(row['SALDO']).replace(',','').replace('$','')) if row['DIAS VENCIDO'] <= 0 else 0, axis=1
        )
        
        print("Saldos y dotación calculados correctamente")
    
    return df

def calcular_vencimientos_historicos(df, fecha_cierre_str=None):
    """Calcula vencimientos históricos de los últimos 6 meses"""
    print("Calculando vencimientos históricos...")
    
    fecha_cierre = obtener_fecha_cierre(fecha_cierre_str)
    
    if 'FECHA VTO_DT' in df.columns and 'SALDO' in df.columns:
        # Configurar locale para nombres de meses en español
        try:
            locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_TIME, 'es_CO.UTF-8')
            except:
                locale.setlocale(locale.LC_TIME, '')
        
        # Calcular vencimientos de los últimos 6 meses
        for i in range(1, 7):
            inicio_mes = fecha_cierre - pd.DateOffset(months=i)
            fin_mes = fecha_cierre - pd.DateOffset(months=i-1)
            
            # Nombre del mes en formato abreviado
            nombre_mes = inicio_mes.strftime('%b-%y').lower()
            
            df[nombre_mes] = df.apply(
                lambda row: float(str(row['SALDO']).replace(',','').replace('$',''))
                if (row['FECHA VTO_DT'] and inicio_mes <= row['FECHA VTO_DT'] < fin_mes) 
                else '-', axis=1
            )
        
        print("Vencimientos históricos calculados correctamente")
    
    return df

def calcular_vencimientos_por_rango(df):
    """Calcula vencimientos por rango de días según especificaciones"""
    print("Calculando vencimientos por rango...")
    
    if 'SALDO' in df.columns and 'DIAS VENCIDO' in df.columns:
        for nombre_col, min_dias, max_dias in VENCIMIENTOS_RANGOS:
            df[nombre_col] = df.apply(
                lambda row: float(str(row['SALDO']).replace(',','').replace('$',''))
                if min_dias <= row['DIAS VENCIDO'] <= max_dias 
                else 0, axis=1
            )
        
        print("Vencimientos por rango calculados correctamente")
    
    return df

def calcular_por_vencer(df, fecha_cierre_str=None):
    """Calcula valores por vencer de los próximos 3 meses"""
    print("Calculando valores por vencer...")
    
    fecha_cierre = obtener_fecha_cierre(fecha_cierre_str)
    
    if 'FECHA VTO_DT' in df.columns and 'SALDO' in df.columns:
        # Por vencer próximos 3 meses
        for i in range(1, 4):
            inicio_mes = fecha_cierre + pd.DateOffset(months=i-1)
            fin_mes = fecha_cierre + pd.DateOffset(months=i)
            
            df[f'Por_Vencer_{i}_meses'] = df.apply(
                lambda row: float(str(row['SALDO']).replace(',','').replace('$',''))
                if (row['FECHA VTO_DT'] and inicio_mes <= row['FECHA VTO_DT'] < fin_mes) 
                else 0, axis=1
            )
        
        # Mayor a 90 días
        fecha_90_dias = fecha_cierre + pd.DateOffset(days=90)
        df['Por_Vencer_+90_dias'] = df.apply(
            lambda row: float(str(row['SALDO']).replace(',','').replace('$',''))
            if (row['FECHA VTO_DT'] and row['FECHA VTO_DT'] >= fecha_90_dias) 
            else 0, axis=1
        )
        
        print("Valores por vencer calculados correctamente")
    
    return df

def validar_saldos(df):
    """Valida que las sumas de saldos sean correctas"""
    print("Validando saldos...")
    
    errores = []
    
    # Validar que Mora Total + Valor Total Por Vencer = Saldo
    if all(col in df.columns for col in ['Mora Total', 'Valor Total Por Vencer', 'SALDO']):
        df['Verificación Suma Saldos'] = df.apply(
            lambda row: 'OK' if abs(row['Mora Total'] + row['Valor Total Por Vencer'] - float(str(row['SALDO']).replace(',','').replace('$',''))) < 0.01
            else 'ERROR', axis=1
        )
        
        errores_suma = (df['Verificación Suma Saldos'] == 'ERROR').sum()
        if errores_suma > 0:
            print(f"ADVERTENCIA: {errores_suma} registros con error en suma de saldos")
            errores.append(f"Suma saldos: {errores_suma} errores")
    
    # Validar que suma de vencimientos = saldo
    columnas_vencimiento = [col for col, _, _ in VENCIMIENTOS_RANGOS]
    if all(col in df.columns for col in columnas_vencimiento) and 'SALDO' in df.columns:
        df['Validación Vencimientos'] = df.apply(
            lambda row: 'OK' if abs(sum([row[col] for col in columnas_vencimiento]) - float(str(row['SALDO']).replace(',','').replace('$',''))) < 0.01
            else 'ERROR', axis=1
        )
        
        errores_venc = (df['Validación Vencimientos'] == 'ERROR').sum()
        if errores_venc > 0:
            print(f"ADVERTENCIA: {errores_venc} registros con error en suma de vencimientos")
            errores.append(f"Vencimientos: {errores_venc} errores")
    
    if errores:
        print(f"Errores encontrados: {', '.join(errores)}")
    else:
        print("Todas las validaciones de saldos son correctas")
    
    return df

def crear_deuda_incobrable(df):
    """Crea la columna de deuda incobrable"""
    print("Creando columna de deuda incobrable...")
    
    if '  Valor Dotación  ' in df.columns:
        df['  DEUDA INCOBRABLE  '] = df['  Valor Dotación  ']
        print("Columna de deuda incobrable creada correctamente")
    
    return df

def aplicar_formato_final(df):
    """Aplica el formato final al DataFrame"""
    print("Aplicando formato final...")
    
    # Eliminar columnas de datetime que contienen información de tiempo
    columnas_a_eliminar = [col for col in df.columns if col.endswith('_DT')]
    if columnas_a_eliminar:
        print(f"Eliminando columnas de datetime: {columnas_a_eliminar}")
        df = df.drop(columns=columnas_a_eliminar)
    
    # Columnas numéricas que requieren formato colombiano
    columnas_numericas = [
        'SALDO', 'SALDO VENCIDO', '  Valor Dotación  ', 'Mora Total', 
        'Valor Total Por Vencer', '  DEUDA INCOBRABLE  '
    ]
    
    # Agregar columnas de vencimientos
    columnas_numericas.extend([col for col, _, _ in VENCIMIENTOS_RANGOS])
    
    # Agregar columnas de vencimientos históricos
    fecha_cierre = obtener_fecha_cierre()
    for i in range(1, 7):
        inicio_mes = fecha_cierre - pd.DateOffset(months=i)
        nombre_mes = inicio_mes.strftime('%b-%y').lower()
        columnas_numericas.append(nombre_mes)
    
    # Agregar columnas por vencer
    for i in range(1, 4):
        columnas_numericas.append(f'Por_Vencer_{i}_meses')
    columnas_numericas.append('Por_Vencer_+90_dias')
    
    # Filtrar solo las columnas que existen en el DataFrame
    columnas_existentes = [col for col in columnas_numericas if col in df.columns]
    
    # Aplicar formato colombiano
    # df = aplicar_formato_colombiano_dataframe(df, columnas_existentes)  # Comentado porque no está definida
    
    # Reemplazar ceros por '-' en columnas numéricas (excepto porcentajes)
    for col in columnas_existentes:
        if '%' not in col:
            df[col] = df[col].replace(['0', '0,00', '0.00', '0,0', '0.0'], '-')
    
    print("Formato final aplicado correctamente")
    return df

def procesar_cartera(input_path, output_path=None, fecha_cierre_str=None):
    """
    Procesa el archivo de cartera según las especificaciones del formato de deuda
    """
    print("=" * 80)
    print("PROCESADOR DE CARTERA - FORMATO DEUDA")
    print("=" * 80)
    
    if fecha_cierre_str:
        print(f"Fecha de cierre especificada: {fecha_cierre_str}")
    else:
        print("Usando fecha de cierre por defecto (último día del mes actual)")
    
    try:
        # Leer archivo CSV
        print(f"Leyendo archivo: {input_path}")
        df = pd.read_csv(input_path, sep=';', encoding='latin1', dtype=str)
        print(f"Archivo leído correctamente. Registros: {len(df)}")
        
        # Procesar datos
        df = limpiar_y_validar_datos(df)
        df = unificar_nombres_clientes(df)
        df = procesar_fechas(df, fecha_cierre_str)
        df = calcular_dias_vencidos(df, fecha_cierre_str)
        df = calcular_saldos_y_dotacion(df)
        df = calcular_vencimientos_historicos(df, fecha_cierre_str)
        df = calcular_vencimientos_por_rango(df)
        df = calcular_por_vencer(df, fecha_cierre_str)
        df = validar_saldos(df)
        df = crear_deuda_incobrable(df)
        df = aplicar_formato_final(df)
        
        # Definir carpeta de salida
        output_dir = r'C:\wamp64\www\cartera\PROVCA_PROCESADOS'
        os.makedirs(output_dir, exist_ok=True)
        
        if not output_path:
            ahora = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            output_path = os.path.join(output_dir, f'CARTERA_PROCESADA_{ahora}.xlsx')
        
        # Verificar que el DataFrame no esté vacío
        if df.empty:
            print("ERROR: El DataFrame está vacío. No se puede generar archivo.")
            return None
        
        # Guardar archivo Excel
        print(f"Guardando archivo: {output_path}")
        df.to_excel(output_path, index=False)
        
        # Verificar que el archivo se creó correctamente
        if not os.path.exists(output_path):
            print("ERROR: No se pudo crear el archivo Excel.")
            return None
        
        if os.path.getsize(output_path) == 0:
            print("ERROR: El archivo Excel está vacío.")
            os.remove(output_path)
            return None
        
        # Ajustar formato de Excel
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Ajustar alineación: números a la derecha, texto al centro
            for row in ws.iter_rows(min_row=2):  # Saltar encabezados
                for cell in row:
                    valor = cell.value
                    if valor is None:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        try:
                            # Si es número
                            float(str(valor).replace('.','').replace(',','').replace('%',''))
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        except:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Encabezados al centro
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            wb.save(output_path)
            print("Formato de Excel ajustado correctamente")
        except Exception as e:
            print(f"Advertencia: No se pudo ajustar el formato de Excel: {e}")
        
        # Resumen final
        print("\n" + "=" * 80)
        print("PROCESAMIENTO COMPLETADO EXITOSAMENTE")
        print("=" * 80)
        print(f"Archivo procesado: {input_path}")
        print(f"Archivo generado: {output_path}")
        print(f"Registros procesados: {len(df)}")
        print(f"Columnas generadas: {len(df.columns)}")
        
        # Mostrar columnas principales
        columnas_principales = [
            'EMPRESA', 'CODIGO CLIENTE', 'NOMBRE', 'DENOMINACION COMERCIAL',
            'NUMERO FACTURA', 'FECHA VTO', 'SALDO', 'DIAS VENCIDO',
            '% Dotación', '  Valor Dotación  ', 'Mora Total', 'Valor Total Por Vencer'
        ]
        print(f"\nColumnas principales: {[col for col in columnas_principales if col in df.columns]}")
        
        return output_path
        
    except Exception as e:
        print(f"ERROR durante el procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        fecha_cierre = sys.argv[2] if len(sys.argv) > 2 else None
        output_file = sys.argv[3] if len(sys.argv) > 3 else None
        procesar_cartera(input_file, output_file, fecha_cierre)
    else:
        print("Uso: python procesador_cartera.py <ruta_entrada_csv> [<fecha_cierre_YYYY-MM-DD>] [<ruta_salida_excel>]")

def procesar_archivo():
    return None