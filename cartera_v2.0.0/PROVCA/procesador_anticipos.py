#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador de Anticipos
Versión corregida compatible con modelo_deuda.py
"""

import pandas as pd
import os
import sys
from datetime import datetime

# -------------------------
# Configuración
# -------------------------

MAPEO_ANTICIPO = {
    'NCCDEM': 'EMPRESA',
    'NCCDAC': 'ACTIVIDAD',
    'NCCDCL': 'CODIGO CLIENTE',
    'WWNIT': 'NIT/CEDULA',
    'WWNMCL': 'NOMBRE COMERCIAL',
    'WWNMDO': 'DIRECCION',
    'WWTLF1': 'TELEFONO',
    'WWNMPO': 'POBLACION',
    'CCCDFB': 'CODIGO AGENTE',
    'BDNMNM': 'NOMBRE AGENTE',
    'BDNMPA': 'APELLIDO AGENTE',
    'NCMOMO': 'TIPO ANTICIPO',
    'NCCDR3': 'NRO ANTICIPO',
    'NCIMAN': 'VALOR ANTICIPO',
    'NCFEGR': 'FECHA ANTICIPO'
}

# -------------------------
# Procesador
# -------------------------

def procesar_anticipos(input_path, output_path=None):
    """Procesa archivo de anticipos para que modelo_deuda.py pueda usarlo"""

    # Leer archivo
    df = pd.read_csv(input_path, sep=';', encoding='latin1', dtype=str)

    # Renombrar columnas según mapeo
    df = df.rename(columns={k: v for k, v in MAPEO_ANTICIPO.items() if k in df.columns})

    # Convertir a numérico y multiplicar por -1
    if 'VALOR ANTICIPO' in df.columns:
        df['VALOR ANTICIPO'] = (
            df['VALOR ANTICIPO']
            .astype(str)
            .str.replace('.', '', regex=False)   # quitar separador de miles
            .str.replace(',', '.', regex=False)  # coma decimal a punto
        )
        df['VALOR ANTICIPO'] = pd.to_numeric(df['VALOR ANTICIPO'], errors='coerce').fillna(0.0)
        df['VALOR ANTICIPO'] *= -1

    # Formatear fecha como string
    if 'FECHA ANTICIPO' in df.columns:
        df['FECHA ANTICIPO'] = pd.to_datetime(df['FECHA ANTICIPO'], errors='coerce').dt.strftime('%d-%m-%Y')

    # Validación
    if df.empty or len(df.columns) == 0:
        print("❌ ERROR: No hay datos para procesar.")
        return

    # 🔹 Asegurar columnas necesarias para modelo_deuda.py (todas como float)
    columnas_necesarias = [
        'SALDO', 'SALDO VENCIDO', 'SALDO NO VENCIDO',
        'VENCIDO 30', 'VENCIDO 60', 'VENCIDO 90',
        'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360',
        'DEUDA INCOBRABLE'
    ]

    # Crear SALDO desde VALOR ANTICIPO
    if 'VALOR ANTICIPO' in df.columns:
        df['SALDO'] = df['VALOR ANTICIPO']
    else:
        df['SALDO'] = 0.0

    # Rellenar otras columnas en cero
    for col in columnas_necesarias:
        if col not in df.columns:
            df[col] = 0.0

    # Convertir todas las columnas numéricas a float
    for col in columnas_necesarias:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

    # Carpeta por defecto (alineada con el resto del proyecto)
    output_dir = r'C:\wamp64\www\modelo-deuda-python\cartera_v2.0.0\PROVCA_PROCESADOS'
    os.makedirs(output_dir, exist_ok=True)

    # Definir ruta de salida
    if not output_path:
        ahora = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        output_path = os.path.join(output_dir, f'ANTICIPO_{ahora}.xlsx')
    elif os.path.isdir(output_path):
        ahora = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        output_path = os.path.join(output_path, f'ANTICIPO_{ahora}.xlsx')
    elif not output_path.lower().endswith(('.xlsx', '.xls')):
        output_path += '.xlsx'

    # Filtrar filas con todos los valores monetarios en cero
    monto_cols = [
        'SALDO', 'SALDO VENCIDO', 'SALDO NO VENCIDO', 'VENCIDO 30', 'VENCIDO 60',
        'VENCIDO 90', 'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360', 'DEUDA INCOBRABLE'
    ]
    presentes = [c for c in monto_cols if c in df.columns]
    if presentes:
        df = df[df[presentes].abs().sum(axis=1) != 0]

    # Guardar archivo con formato de ceros como '-'
    try:
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Anticipos')
            ws = writer.sheets['Anticipos']
            book = writer.book
            fmt = book.add_format({"num_format": "#,##0;-#,##0;\"-\";@"})
            for col in presentes:
                if col in df.columns:
                    idx = df.columns.get_loc(col)
                    ws.set_column(idx, idx, None, fmt)
    except Exception:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Anticipos')
            from openpyxl.utils import get_column_letter
            ws = writer.sheets['Anticipos']
            fmt = "#,##0;-#,##0;\"-\";@"
            for col in presentes:
                if col in df.columns:
                    j = df.columns.get_loc(col) + 1
                    letter = get_column_letter(j)
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f"{letter}{row}"]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = fmt

    print(f"\n✅ Archivo de anticipos procesado y guardado en: {output_path}")
    print(f"📊 Registros procesados: {len(df)}")
    print(f"🗂 Columnas generadas: {len(df.columns)}")

# -------------------------
# Menú interactivo
# -------------------------

def menu():
    while True:
        print("\n=== Menú de Procesamiento de Anticipos ===")
        print("1. Procesar archivo de anticipos")
        print("2. Salir")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            input_file = input("Ingrese la ruta del archivo de anticipos: ").strip('"')
            output_file = input("Ingrese la ruta de salida (opcional): ").strip('"')
            output_file = output_file if output_file else None
            try:
                procesar_anticipos(input_file, output_file)
            except Exception as e:
                print(f"\n❌ Error: {e}")
        elif opcion == "2":
            print("👋 Saliendo del programa...")
            sys.exit(0)
        else:
            print("⚠ Opción no válida, intente de nuevo.")

# -------------------------
# Ejecución
# -------------------------

if __name__ == "__main__":
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
        procesar_anticipos(input_file, output_file)
    else:
        menu()
