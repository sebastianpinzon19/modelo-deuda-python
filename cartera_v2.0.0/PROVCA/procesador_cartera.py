# -*- coding: utf-8 -*-
"""
Procesador de Cartera PROVCA
- Conserva funciones originales
- Corrige generación de meses históricos desde fecha de cierre exacta
- Llena DENOMINACION COMERCIAL con NOMBRE si está vacío
- Mantiene columnas de día, mes y año de fechas
"""

import pandas as pd
from datetime import datetime
import os
import sys
from trm_config import load_trm, save_trm, parse_trm_value, format_trm_display

# -------------------
# Configuración
# -------------------
RENOMBRES = {
    "PCCDEM": "EMPRESA",
    "PCCDAC": "ACTIVIDAD",
    "PCDEAC": "EMPRESA_2",
    "PCCDAG": "CODIGO AGENTE",
    "PCNMAG": "AGENTE",
    "PCCDCO": "CODIGO COBRADOR",
    "PCNMCO": "COBRADOR",
    "PCCDCL": "CODIGO CLIENTE",
    "PCCDDN": "IDENTIFICACION",
    "PCNMCL": "NOMBRE",
    "PCNMCM": "DENOMINACION COMERCIAL",
    "PCNMDO": "DIRECCION",
    "PCTLF1": "TELEFONO",
    "PCNMPO": "CIUDAD",
    "PCNUFC": "NUMERO FACTURA",
    "PCORPD": "TIPO",
    "PCFEFA": "FECHA",
    "PCFEVE": "FECHA VTO",
    "PCVAFA": "VALOR",
    "PCSALD": "SALDO"
}

VENCIMIENTOS_RANGOS = [
    ("SALDO NO VENCIDO", 0, 29),
    ("VENCIDO 30", 30, 59),
    ("VENCIDO 60", 60, 89),
    ("VENCIDO 90", 90, 179),
    ("VENCIDO 180", 180, 359),
    ("VENCIDO 360", 360, 369),
    ("VENCIDO + 360", 370, 999999),
]

OUT_DIR = r"C:\wamp64\www\modelo-deuda-python\cartera_v2.0.0\PROVCA_PROCESADOS"
os.makedirs(OUT_DIR, exist_ok=True)

# -------------------
# Helpers
# -------------------
def obtener_fecha_cierre(fecha_cierre_str=None):
    if fecha_cierre_str:
        dt = pd.to_datetime(fecha_cierre_str, errors="coerce")
        if pd.isna(dt):
            raise ValueError("Fecha de cierre inválida. Use YYYY-MM-DD")
        return pd.to_datetime(dt).normalize()
    return pd.to_datetime(datetime.today()).normalize()

def convertir_valor_to_float(x):
    if pd.isna(x):
        return 0.0
    s = str(x).strip()
    if s == "" or s in ["-", "0"]:
        return 0.0
    s = s.replace("$", "").replace(" ", "").replace("\u200b", "")
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        digits = "".join(ch for ch in s if ch.isdigit() or ch == ".")
        return float(digits) if digits else 0.0

def formato_fecha_str(dt):
    if pd.isna(dt):
        return ""
    return pd.to_datetime(dt).strftime("%d/%m/%Y")

def parse_fecha_segura(serie):
    def try_parse(val):
        if pd.isna(val) or str(val).strip() == "":
            return pd.NaT
        formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%Y%m%d"]
        for fmt in formatos:
            try:
                return datetime.strptime(str(val).strip(), fmt)
            except ValueError:
                continue
        return pd.to_datetime(val, errors="coerce", dayfirst=True)
    return serie.apply(try_parse)

# -------------------
# Escritura con formato de miles en Excel
# -------------------
def guardar_excel_con_miles(df, output_path, columnas_montos, aplicar_formato=True):
    # Quitar la columna LINEA_CLAVE si existe
    if 'LINEA_CLAVE' in df.columns:
        df = df.drop(columns=['LINEA_CLAVE'])
    
    # Asegurar que las columnas de montos sean numéricas
    for col in columnas_montos:
        if col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col], errors='coerce')
            except Exception:
                pass

    try:
        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Cartera")
            if aplicar_formato:
                ws = writer.sheets["Cartera"]
                book = writer.book
                
                # Aplicar formato de miles sin decimales a todas las columnas de montos
                fmt_miles = book.add_format({"num_format": "#,##0;-#,##0;\"-\";@"})
                for col in columnas_montos:
                    if col in df.columns:
                        idx = df.columns.get_loc(col)
                        ws.set_column(idx, idx, 18, fmt_miles)
        return
    except Exception as e:
        print(f"Error con xlsxwriter: {e}")
        pass

    # Fallback con openpyxl
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Cartera")
        if aplicar_formato:
            from openpyxl.utils import get_column_letter
            ws = writer.sheets["Cartera"]
            
            for col in columnas_montos:
                if col in df.columns:
                    j = df.columns.get_loc(col) + 1
                    letter = get_column_letter(j)
                    
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f"{letter}{row}"]
                        if isinstance(cell.value, (int, float)) and cell.value != 0:
                            # Forzar formato con miles sin decimales en todos los casos
                            cell.number_format = "#,##0;-#,##0;\"-\";@"

# -------------------
# Procesador principal
# -------------------
def procesar_cartera(input_path, output_path=None, fecha_cierre_str=None, formatear_miles_excel=True, override_moneda=None):
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"No se encontró el archivo CSV: {input_path}")
    df = pd.read_csv(input_path, sep=";", encoding="latin1", dtype=str, keep_default_na=False, na_values=[""])
    df.rename(columns=RENOMBRES, inplace=True)
    if "PCIMCO" in df.columns:
        df.drop(columns=["PCIMCO"], inplace=True)

    if "MONEDA" not in df.columns:
        df["MONEDA"] = "PESOS COL"
    if override_moneda and isinstance(override_moneda, str) and override_moneda.strip() != "":
        df["MONEDA"] = override_moneda.strip()

    df["SALDO"] = df.get("SALDO", 0).apply(convertir_valor_to_float)
    df["VALOR"] = df.get("VALOR", 0).apply(convertir_valor_to_float)

    # Rellenar DENOMINACION COMERCIAL con NOMBRE si está vacío (robusto ante espacios y \u200b)
    if "DENOMINACION COMERCIAL" in df.columns and "NOMBRE" in df.columns:
        df["DENOMINACION COMERCIAL"] = (
            df["DENOMINACION COMERCIAL"].astype(str).str.replace("\u200b", "", regex=False).str.strip()
        )
        df["NOMBRE"] = (
            df["NOMBRE"].astype(str).str.replace("\u200b", "", regex=False).str.strip()
        )
        mask_dc_vacia = df["DENOMINACION COMERCIAL"].isna() | (df["DENOMINACION COMERCIAL"].str.strip() == "")
        mask_nombre_ok = df["NOMBRE"].notna() & (df["NOMBRE"].str.strip() != "")
        df.loc[mask_dc_vacia & mask_nombre_ok, "DENOMINACION COMERCIAL"] = df.loc[mask_dc_vacia & mask_nombre_ok, "NOMBRE"]

    for col in ["FECHA", "FECHA VTO"]:
        if col in df.columns:
            df[col] = parse_fecha_segura(df[col])

    fecha_cierre = obtener_fecha_cierre(fecha_cierre_str)
    if "FECHA VTO" in df.columns:
        df["DIAS VENCIDO"] = df["FECHA VTO"].apply(lambda fv: max((fecha_cierre - fv).days, 0) if pd.notna(fv) else 0).astype(int)
        df["DIAS POR VENCER"] = df["FECHA VTO"].apply(lambda fv: max((fv - fecha_cierre).days, 0) if pd.notna(fv) else 0).astype(int)
    else:
        df["DIAS VENCIDO"] = 0
        df["DIAS POR VENCER"] = 0

    df["SALDO VENCIDO"] = df.apply(lambda r: r["SALDO"] if r["DIAS VENCIDO"] > 0 else 0.0, axis=1)
    df["% Dotación"] = df["DIAS VENCIDO"].apply(lambda x: "100%" if x >= 180 else "0%")
    df["Valor Dotación"] = df.apply(lambda r: r["SALDO"] if r["DIAS VENCIDO"] >= 180 else 0.0, axis=1)
    df["Mora Total"] = df["SALDO VENCIDO"]
    df["Valor Total Por Vencer"] = df.apply(lambda r: r["SALDO"] if r["DIAS VENCIDO"] == 0 else 0.0, axis=1)
    df[">=180 días"] = df.apply(lambda r: r["SALDO"] if r["DIAS VENCIDO"] >= 180 else 0.0, axis=1)

    # Meses históricos desde la fecha de cierre
    meses_hist = []
    for i in range(0, 6):
        mes_fecha = (fecha_cierre - pd.DateOffset(months=i)).replace(day=1)
        siguiente_mes = mes_fecha + pd.DateOffset(months=1)
        col = mes_fecha.strftime("%b-%y").lower()
        meses_hist.append(col)
        df[col] = df.apply(lambda r, a=mes_fecha, b=siguiente_mes: r["SALDO"] if (pd.notna(r["FECHA VTO"]) and (a <= r["FECHA VTO"] < b)) else 0.0, axis=1)

    # Por vencer
    por_vencer_cols = []
    for i in range(1, 4):
        a = fecha_cierre + pd.DateOffset(months=i)
        b = fecha_cierre + pd.DateOffset(months=i + 1)
        col = f"Por_Vencer_{i}_meses"
        por_vencer_cols.append(col)
        df[col] = df.apply(lambda r, a=a, b=b: r["SALDO"] if (pd.notna(r["FECHA VTO"]) and (a <= r["FECHA VTO"] < b)) else 0.0, axis=1)

    fecha_90 = fecha_cierre + pd.DateOffset(days=90)
    df["Por_Vencer_+90_dias"] = df.apply(lambda r: r["SALDO"] if (pd.notna(r["FECHA VTO"]) and r["FECHA VTO"] >= fecha_90) else 0.0, axis=1)

    for name, mind, maxd in VENCIMIENTOS_RANGOS:
        df[name] = df.apply(lambda r, mind=mind, maxd=maxd: r["SALDO"] if (mind <= r["DIAS VENCIDO"] <= maxd) else 0.0, axis=1)

    df["DEUDA INCOBRABLE"] = df["Valor Dotación"]

    if "FECHA" in df.columns:
        df["DIA FECHA"] = df["FECHA"].dt.day.fillna(0).astype(int)
        df["MES FECHA"] = df["FECHA"].dt.month.fillna(0).astype(int)
        df["AÑO FECHA"] = df["FECHA"].dt.year.fillna(0).astype(int)
        df["FECHA"] = df["FECHA"].apply(formato_fecha_str)
    if "FECHA VTO" in df.columns:
        df["DIA FECHA VTO"] = df["FECHA VTO"].dt.day.fillna(0).astype(int)
        df["MES FECHA VTO"] = df["FECHA VTO"].dt.month.fillna(0).astype(int)
        df["AÑO FECHA VTO"] = df["FECHA VTO"].dt.year.fillna(0).astype(int)
        df["FECHA VTO"] = df["FECHA VTO"].apply(formato_fecha_str)

    base_cols = [
        "EMPRESA", "ACTIVIDAD", "EMPRESA_2", "CODIGO AGENTE", "AGENTE",
        "CODIGO COBRADOR", "COBRADOR", "CODIGO CLIENTE", "IDENTIFICACION",
        "NOMBRE", "DENOMINACION COMERCIAL", "DIRECCION", "TELEFONO", "CIUDAD",
        "NUMERO FACTURA", "TIPO", "MONEDA"
    ]
    fechas_cols = []
    if "FECHA" in df.columns:
        fechas_cols += ["FECHA", "DIA FECHA", "MES FECHA", "AÑO FECHA"]
    if "FECHA VTO" in df.columns:
        fechas_cols += ["FECHA VTO", "DIA FECHA VTO", "MES FECHA VTO", "AÑO FECHA VTO"]

    fin_cols = ["VALOR", "SALDO", "DIAS VENCIDO", "DIAS POR VENCER", "SALDO VENCIDO",
                "% Dotación", "Valor Dotación", "Mora Total", "Valor Total Por Vencer", ">=180 días"]

    final_cols = base_cols + fechas_cols + fin_cols + meses_hist + por_vencer_cols + ["Por_Vencer_+90_dias"] + [v[0] for v in VENCIMIENTOS_RANGOS] + ["DEUDA INCOBRABLE"]

    df = df[[c for c in final_cols if c in df.columns]]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(os.path.basename(input_path))[0]
    if not output_path:
        output_path = os.path.join(OUT_DIR, f"{base_name}_cartera_{timestamp}.xlsx")
    elif os.path.isdir(output_path):
        output_path = os.path.join(output_path, f"{base_name}_cartera_{timestamp}.xlsx")

    columnas_montos = ["VALOR", "SALDO", "SALDO VENCIDO", "Valor Dotación", "Mora Total",
                       "Valor Total Por Vencer", "DEUDA INCOBRABLE", ">=180 días"] + meses_hist + por_vencer_cols + ["Por_Vencer_+90_dias"] + [v[0] for v in VENCIMIENTOS_RANGOS]

    # Eliminar filas sin información monetaria (todos los montos en 0)
    monto_cols_presentes = [c for c in columnas_montos if c in df.columns]
    if monto_cols_presentes:
        df = df[df[monto_cols_presentes].abs().sum(axis=1) != 0]

    guardar_excel_con_miles(df, output_path, columnas_montos, aplicar_formato=formatear_miles_excel)
    print(f"✅ Archivo generado: {output_path}")
    return output_path

# -------------------
# CLI
# -------------------
def menu():
    print("=== Menú de Procesamiento de Cartera ===")
    input_path = input("Ruta del archivo CSV: ").strip()
    output_path = input("Ruta de salida (opcional): ").strip() or None
    fecha_cierre_str = input("Fecha de cierre EXACTA (YYYY-MM-DD, vacío = hoy): ").strip() or None
    cfg = load_trm()
    print(f"TRM guardadas → USD: {format_trm_display(cfg.get('usd'))} | EUR: {format_trm_display(cfg.get('eur'))} (actualizado: {cfg.get('updated_at')})")
    try:
        actualizar = input("¿Desea actualizar TRM para que la usen los demás procesos? (s/n): ").strip().lower()
    except Exception:
        actualizar = 'n'
    if actualizar == 's':
        usd_txt = input("TRM Dólar (deje vacío para mantener): ").strip()
        eur_txt = input("TRM Euro (deje vacío para mantener): ").strip()
        try:
            new_usd = parse_trm_value(usd_txt, cfg.get('usd'))
            new_eur = parse_trm_value(eur_txt, cfg.get('eur'))
            save_trm(new_usd, new_eur)
            print(f"✅ TRM guardadas. USD: {format_trm_display(new_usd)} | EUR: {format_trm_display(new_eur)}")
        except Exception as e:
            print(f"⚠ No se pudieron guardar las TRM: {e}")
    try:
        # Siempre aplicar formato de miles en Excel
        ruta_salida = procesar_cartera(input_path, output_path, fecha_cierre_str, True)
        print("Procesamiento completado exitosamente")
        print(f"Ruta de salida: {ruta_salida}")
    except Exception as e:
        print("ERROR:", e)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        try:
            procesar_cartera(
                sys.argv[1],
                sys.argv[2] if len(sys.argv) > 2 else None,
                sys.argv[3] if len(sys.argv) > 3 else None,
                ((len(sys.argv) > 4) and (str(sys.argv[4]).strip().lower() == "s")) or True,
                sys.argv[5] if len(sys.argv) > 5 else None,
            )
            # Si se pasaron TRM por CLI, guardarlas para uso global
            if len(sys.argv) > 6:
                try:
                    usd_cli = float(str(sys.argv[5]).replace(',', '.'))
                    eur_cli = float(str(sys.argv[6]).replace(',', '.'))
                    save_trm(usd_cli, eur_cli)
                    print(f"ℹ TRM guardadas desde CLI. USD: {usd_cli} | EUR: {eur_cli}")
                except Exception as e:
                    print(f"⚠ TRM por CLI inválidas: {e}")
        except Exception as e:
            print("ERROR:", e)
    else:
        menu()
