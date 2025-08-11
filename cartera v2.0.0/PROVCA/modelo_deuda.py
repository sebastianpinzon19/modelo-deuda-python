# modelo_deuda.py
# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
import os
import sys
from datetime import datetime
from utilidades import convertir_valor, formatear_numero_colombiano, aplicar_formato_colombiano_dataframe
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import gc

# Configurar codificación para evitar problemas con caracteres especiales
import locale
try:
    locale.setlocale(locale.LC_ALL, 'es_CO.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Spanish_Colombia.1252')
    except:
        pass

# --- Configuración de líneas de venta según especificaciones exactas ---
LINEAS_PESOS = [
    ("CT", 80), ("ED", 41), ("ED", 44), ("ED", 47),
    ("PL", 10), ("PL", 15), ("PL", 20), ("PL", 21), ("PL", 23), ("PL", 25), ("PL", 28), ("PL", 29), ("PL", 31), ("PL", 32), ("PL", 53), ("PL", 56), ("PL", 60), ("PL", 62), ("PL", 63), ("PL", 64), ("PL", 65), ("PL", 66), ("PL", 69)
]
LINEAS_DIVISAS = [
    ("PL", 11), ("PL", 18), ("PL", 57), ("PL", 41)
]

# --- Tabla de correspondencia Negocio/Canal según especificaciones exactas ---
TABLA_CANAL = {
    "PL15": ("E-COMMERCE", "PL15"),
    "PL20": ("LIBRERIAS 1", "PL20"),
    "PL25": ("LIBRERIAS 1", "PL25"),
    "PL10": ("LIBRERIAS 2", "PL10"),
    "PL21": ("LIBRERIAS 2", "PL21"),
    "PL53": ("LIBRERIAS 3", "PL53"),
    "PL63": ("LIBRERIAS 3", "PL63"),
    "PL66": ("OTOS DIGITAL", "PL66"),
    "PL60": ("OTROS", "PL60"),
    "PL64": ("OTROS", "PL64"),
    "PL65": ("OTROS", "PL65"),
    "PL28": ("SALDOS", "PL28"),
    "PL29": ("SALDOS", "PL29"),
    "PL31": ("SALDOS", "PL31"),
    "PL18": ("EXPORTACION", "PL18"),
    "PL11": ("EXPORTACION", "PL11"),
    "PL57": ("AULA", "PL57"),
    "PL41": ("OTR", "PL41"),
    "CT80": ("TINTA CLUB DEL LIBRO", "CT80")
}



# --- Función para pedir archivo ---
def pedir_archivo(mensaje, carpeta):
    archivos = [f for f in os.listdir(carpeta) if f.endswith('.xlsx')]
    print(f"Archivos disponibles en {carpeta}:")
    for i, f in enumerate(archivos):
        print(f"{i+1}. {f}")
    idx = int(input(mensaje)) - 1
    return os.path.join(carpeta, archivos[idx])

# --- Función para guardar Excel con múltiples hojas ---
def guardar_excel_con_hojas(df_pesos, df_divisas, df_vencimiento, ruta_salida):
    """Guarda el archivo Excel con las tres hojas: PESOS, DIVISAS, VENCIMIENTO con formato contable"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear workbook nuevo
        wb = Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Definir estilos
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(size=10)
        text_alignment = Alignment(horizontal="left", vertical="center")
        number_alignment = Alignment(horizontal="right", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Función para aplicar formato contable a una celda
        def aplicar_formato_contable(cell, valor, es_porcentaje=False):
            """Aplica formato contable a una celda"""
            try:
                # Limpiar el valor
                if pd.isna(valor) or valor == '' or str(valor).strip() == '':
                    cell.value = '-'
                    cell.alignment = number_alignment
                    cell.font = data_font
                    cell.border = border
                    return
                
                # Convertir a número si es posible
                if isinstance(valor, str):
                    # Remover formato colombiano para convertir
                    valor_limpio = str(valor).replace('.', '').replace(',', '.')
                    try:
                        valor_num = float(valor_limpio)
                    except:
                        # Si no es número, mantener como texto
                        cell.value = str(valor)
                        cell.alignment = text_alignment
                        cell.font = data_font
                        cell.border = border
                        return
                else:
                    valor_num = float(valor)
                
                # Aplicar formato según el tipo
                if es_porcentaje:
                    if valor_num == 0:
                        cell.value = '-'
                    else:
                        cell.value = f"{valor_num:.2f}%"
                    cell.number_format = '0.00%'
                else:
                    if valor_num == 0:
                        cell.value = '-'
                    else:
                        # Formato contable colombiano
                        cell.value = valor_num
                        cell.number_format = '#,##0.00'
                
                cell.alignment = number_alignment
                cell.font = data_font
                cell.border = border
                
            except Exception as e:
                print(f"Error aplicando formato a celda: {e}")
                cell.value = str(valor)
                cell.alignment = text_alignment
                cell.font = data_font
                cell.border = border
        
        # Función para procesar una hoja
        def procesar_hoja(ws, df, nombre_hoja):
            """Procesa una hoja con formato contable"""
            # Escribir encabezados
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Escribir datos con formato
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                for col_idx, col_name in enumerate(df.columns, 1):
                    valor = row[col_name]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Determinar si es columna numérica
                    es_porcentaje = 'dotación' in col_name.lower() or '%' in col_name.lower()
                    es_numerico = any(palabra in col_name.lower() for palabra in [
                        'saldo', 'valor', 'vencido', 'deuda', 'anticipo', 'mora', 'total'
                    ])
                    
                    if es_numerico:
                        aplicar_formato_contable(cell, valor, es_porcentaje)
                    else:
                        # Columna de texto
                        cell.value = str(valor) if not pd.isna(valor) else ''
                        cell.alignment = text_alignment
                        cell.font = data_font
                        cell.border = border
            
            # Ajustar ancho de columnas
            for col_idx, col_name in enumerate(df.columns, 1):
                max_length = max(
                    len(str(col_name)),
                    df[col_name].astype(str).str.len().max() if not df[col_name].empty else 0
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Procesar cada hoja
        if not df_pesos.empty:
            ws_pesos = wb.create_sheet("PESOS")
            procesar_hoja(ws_pesos, df_pesos, "PESOS")
        
        if not df_divisas.empty:
            ws_divisas = wb.create_sheet("DIVISAS")
            procesar_hoja(ws_divisas, df_divisas, "DIVISAS")
        
        if not df_vencimiento.empty:
            ws_vencimiento = wb.create_sheet("VENCIMIENTO")
            procesar_hoja(ws_vencimiento, df_vencimiento, "VENCIMIENTO")
        
        # Asegurar que el directorio existe
        os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
        
        # Guardar archivo
        wb.save(ruta_salida)
        wb.close()
        
        # Verificar que el archivo se creó correctamente
        if os.path.exists(ruta_salida) and os.path.getsize(ruta_salida) > 0:
            print(f"Archivo guardado exitosamente: {ruta_salida}")
            print(f"Tamaño del archivo: {os.path.getsize(ruta_salida)} bytes")
            return True
        else:
            print(f"Error: El archivo no se creó correctamente o está vacío")
            return False
            
    except PermissionError as e:
        print(f"Error de permisos al guardar: {e}")
        return False
    except Exception as e:
        print(f"Error inesperado al guardar: {e}")
        return False
    finally:
        # Limpiar memoria
        gc.collect()

# --- Función para guardar Excel de forma segura ---
def guardar_excel_seguro(df, ruta_salida):
    """Guarda el DataFrame en Excel de forma segura con manejo de errores"""
    try:
        # Crear workbook nuevo
        wb = Workbook()
        ws = wb.active
        ws.title = "Modelo Deuda"
        
        # Convertir DataFrame a valores simples (sin fórmulas)
        df_clean = df.copy()
        
        # Asegurar que todos los valores sean strings o números simples
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(str)
        
        # Escribir encabezados
        headers = list(df_clean.columns)
        ws.append(headers)
        
        # Escribir datos fila por fila (método más seguro)
        for index, row in df_clean.iterrows():
            row_data = []
            for col in headers:
                value = row[col]
                # Convertir valores especiales
                if pd.isna(value) or value == 'nan' or value == 'None':
                    row_data.append('')
                elif isinstance(value, (int, float)):
                    row_data.append(float(value))
                else:
                    row_data.append(str(value))
            ws.append(row_data)
        
        # Asegurar que el directorio existe
        os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
        
        # Guardar con manejo de errores
        wb.save(ruta_salida)
        wb.close()
        
        # Verificar que el archivo se creó correctamente
        if os.path.exists(ruta_salida) and os.path.getsize(ruta_salida) > 0:
            print(f"Archivo guardado exitosamente: {ruta_salida}")
            print(f"Tamaño del archivo: {os.path.getsize(ruta_salida)} bytes")
            return True
        else:
            print(f"Error: El archivo no se creó correctamente o está vacío")
            return False
            
    except PermissionError as e:
        print(f"Error de permisos al guardar: {e}")
        return False
    except Exception as e:
        print(f"Error inesperado al guardar: {e}")
        return False
    finally:
        # Limpiar memoria
        gc.collect()

# --- Función para crear hoja de vencimiento ---
def crear_hoja_vencimiento(df_cartera):
    """Crea la hoja de vencimiento con totales por cliente según especificaciones exactas"""
    
    # Agrupar por cliente y sumar columnas numéricas
    columnas_agrupar = ['DENOMINACION COMERCIAL', 'ACTIVIDAD', 'MONEDA']
    columnas_sumar = [
        'SALDO', 'SALDO NO VENCIDO', 'VENCIDO 30', 'VENCIDO 60', 
        'VENCIDO 90', 'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360', '  DEUDA INCOBRABLE  '
    ]
    
    # Renombrar SALDO a SALDO TOTAL para coincidir con el cuadro especificado
    if 'SALDO' in df_cartera.columns:
        df_cartera = df_cartera.rename(columns={'SALDO': 'SALDO TOTAL'})
        columnas_sumar[0] = 'SALDO TOTAL'
    
    # Verificar columnas existentes
    columnas_agrupar_existentes = [col for col in columnas_agrupar if col in df_cartera.columns]
    columnas_sumar_existentes = [col for col in columnas_sumar if col in df_cartera.columns]
    
    if not columnas_agrupar_existentes or not columnas_sumar_existentes:
        print("ADVERTENCIA: No se pueden crear totales de vencimiento - faltan columnas")
        return pd.DataFrame()
    
    df_vencimiento = df_cartera.groupby(columnas_agrupar_existentes)[columnas_sumar_existentes].sum().reset_index()
    
    # Agregar columnas según especificaciones exactas
    df_vencimiento['PAIS'] = 'Colombia'
    df_vencimiento['COBRO/PAGO'] = 'CLIENTE'
    df_vencimiento['CLIENTE'] = df_vencimiento['DENOMINACION COMERCIAL']
    
    # Determinar NEGOCIO y CANAL según tabla exacta
    def obtener_negocio_canal(row):
        actividad = str(row['ACTIVIDAD'])
        if actividad in TABLA_CANAL:
            return TABLA_CANAL[actividad]
        else:
            return ('OTROS', 'OTROS')
    
    df_vencimiento[['NEGOCIO', 'CANAL']] = df_vencimiento.apply(
        lambda row: pd.Series(obtener_negocio_canal(row)), axis=1
    )
    
    # Reordenar columnas según el cuadro especificado exactamente
    columnas_finales = [
        'PAIS', 'NEGOCIO', 'CANAL', 'COBRO/PAGO', 'MONEDA', 'CLIENTE', 'SALDO TOTAL'
    ] + [col for col in columnas_sumar_existentes if col != 'SALDO TOTAL']
    
    # Asegurar que todas las columnas existan
    for col in columnas_finales:
        if col not in df_vencimiento.columns:
            df_vencimiento[col] = ''
    
    df_vencimiento = df_vencimiento[columnas_finales]
    
    # Ordenar por moneda y cliente
    df_vencimiento = df_vencimiento.sort_values(['MONEDA', 'CLIENTE'])
    
    return df_vencimiento

# --- Función para validar datos según reglas del modelo de deuda ---
def validar_datos_modelo_deuda(df_cartera, df_anticipos):
    """Valida que los datos cumplan con las reglas del modelo de deuda"""
    errores = []
    advertencias = []
    
    # Validar que existan las columnas requeridas
    columnas_requeridas_cartera = ['EMPRESA', 'ACTIVIDAD', 'DENOMINACION COMERCIAL', 'SALDO', 'MONEDA']
    columnas_requeridas_anticipos = ['EMPRESA', 'ACTIVIDAD', 'NOMBRE COMERCIAL', 'VALOR ANTICIPO']
    
    for col in columnas_requeridas_cartera:
        if col not in df_cartera.columns:
            errores.append(f"Columna requerida '{col}' no encontrada en archivo de cartera")
    
    for col in columnas_requeridas_anticipos:
        if col not in df_anticipos.columns:
            errores.append(f"Columna requerida '{col}' no encontrada en archivo de anticipos")
    
    # Validar líneas de venta
    lineas_pesos = ['80', '41', '44', '47', '10', '15', '20', '21', '23', '25', '28', '29', '31', '32', '53', '56', '60', '62', '63', '64', '65', '66', '69']
    lineas_divisas = ['11', '18', '57', '41']
    
    if 'ACTIVIDAD' in df_cartera.columns:
        actividades_cartera = df_cartera['ACTIVIDAD'].astype(str).unique()
        lineas_no_encontradas = []
        
        for linea in lineas_pesos + lineas_divisas:
            if linea not in actividades_cartera:
                lineas_no_encontradas.append(linea)
        
        if lineas_no_encontradas:
            advertencias.append(f"Líneas de venta no encontradas: {', '.join(lineas_no_encontradas)}")
    
    # Validar monedas
    if 'MONEDA' in df_cartera.columns:
        monedas_validas = ['PESOS COL', 'DOLAR', 'EURO', 'DÓLAR']
        monedas_cartera = df_cartera['MONEDA'].astype(str).unique()
        monedas_invalidas = [m for m in monedas_cartera if m not in monedas_validas and m != 'nan']
        
        if monedas_invalidas:
            advertencias.append(f"Monedas no reconocidas: {', '.join(monedas_invalidas)}")
    
    # Validar valores numéricos
    columnas_numericas = ['SALDO', 'SALDO VENCIDO', 'SALDO NO VENCIDO']
    for col in columnas_numericas:
        if col in df_cartera.columns:
            valores_negativos = df_cartera[df_cartera[col] < 0]
            if not valores_negativos.empty:
                advertencias.append(f"Se encontraron {len(valores_negativos)} registros con valores negativos en '{col}'")
    
    # Validar que no haya facturas del mes siguiente al cierre
    if 'FECHA' in df_cartera.columns:
        from datetime import datetime
        import pandas as pd
        
        # Fecha por defecto: último día del mes actual
        hoy = datetime.now()
        if hoy.month == 12:
            fecha_cierre = datetime(hoy.year + 1, 1, 1) - pd.Timedelta(days=1)
        else:
            fecha_cierre = datetime(hoy.year, hoy.month + 1, 1) - pd.Timedelta(days=1)
        
        # Calcular el primer día del mes siguiente al cierre
        if fecha_cierre.month == 12:
            primer_dia_mes_siguiente = datetime(fecha_cierre.year + 1, 1, 1)
        else:
            primer_dia_mes_siguiente = datetime(fecha_cierre.year, fecha_cierre.month + 1, 1)
        
        # Convertir fechas de factura a datetime
        fechas_factura = []
        for fecha_str in df_cartera['FECHA']:
            try:
                if pd.notna(fecha_str) and str(fecha_str).strip() != '':
                    # Intentar diferentes formatos de fecha
                    fecha_limpia = str(fecha_str).strip()
                    if '/' in fecha_limpia:
                        # Formato dd/mm/yyyy
                        dia, mes, anio = fecha_limpia.split('/')
                        fecha_dt = datetime(int(anio), int(mes), int(dia))
                    elif '-' in fecha_limpia:
                        # Formato dd-mm-yyyy
                        dia, mes, anio = fecha_limpia.split('-')
                        fecha_dt = datetime(int(anio), int(mes), int(dia))
                    else:
                        # Formato yyyymmdd
                        fecha_dt = datetime.strptime(fecha_limpia, "%Y%m%d")
                    fechas_factura.append(fecha_dt)
                else:
                    fechas_factura.append(None)
            except Exception:
                fechas_factura.append(None)
        
        # Verificar si hay facturas del mes siguiente
        facturas_mes_siguiente = []
        for i, fecha_factura in enumerate(fechas_factura):
            if fecha_factura and fecha_factura >= primer_dia_mes_siguiente:
                facturas_mes_siguiente.append(i)
        
        if facturas_mes_siguiente:
            mes_siguiente_nombre = primer_dia_mes_siguiente.strftime("%B %Y")
            errores.append(f"❌ ERROR CRÍTICO: Se encontraron {len(facturas_mes_siguiente)} facturas del mes siguiente ({mes_siguiente_nombre}) al cierre. El archivo de cierre de {fecha_cierre.strftime('%B %Y')} no puede contener facturas de {mes_siguiente_nombre}.")
            print(f"❌ ERROR CRÍTICO: Se encontraron {len(facturas_mes_siguiente)} facturas del mes siguiente ({mes_siguiente_nombre}) al cierre.")
            print(f"   El archivo de cierre de {fecha_cierre.strftime('%B %Y')} no puede contener facturas de {mes_siguiente_nombre}.")
            print(f"   Por favor, revise y corrija las fechas de las facturas antes de continuar.")
    
    # Mostrar resultados
    if errores:
        print("❌ ERRORES ENCONTRADOS:")
        for error in errores:
            print(f"  - {error}")
    
    if advertencias:
        print("⚠️ ADVERTENCIAS:")
        for advertencia in advertencias:
            print(f"  - {advertencia}")
    
    if not errores and not advertencias:
        print("VALIDACION EXITOSA: Todos los datos cumplen con las reglas del modelo de deuda")
    
    return len(errores) == 0

# --- Función para crear totales por moneda ---
def crear_totales_por_moneda(df_vencimiento):
    """Crea totales por moneda al final de la hoja vencimiento según especificaciones exactas"""
    if df_vencimiento.empty:
        return df_vencimiento
    
    totales = []
    columnas_sumar = [
        'SALDO TOTAL', 'SALDO NO VENCIDO', 'VENCIDO 30', 'VENCIDO 60', 
        'VENCIDO 90', 'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360', '  DEUDA INCOBRABLE  '
    ]
    
    # Filtrar columnas existentes
    columnas_sumar_existentes = [col for col in columnas_sumar if col in df_vencimiento.columns]
    
    # Mapeo de monedas según el cuadro especificado exactamente
    mapeo_monedas = {
        'PESOS COL': 'Moneda Local',
        'DOLAR': 'Dólar',
        'EURO': 'Euro'
    }
    
    # Crear totales por moneda específica
    for moneda_original in df_vencimiento['MONEDA'].unique():
        df_moneda = df_vencimiento[df_vencimiento['MONEDA'] == moneda_original]
        
        # Usar el mapeo para el nombre del total
        moneda_total = mapeo_monedas.get(moneda_original, moneda_original)
        
        total = {
            'PAIS': '',
            'NEGOCIO': '',
            'CANAL': '',
            'COBRO/PAGO': '',
            'MONEDA': moneda_total,
            'CLIENTE': ''
        }
        
        for col in columnas_sumar_existentes:
            total[col] = df_moneda[col].sum()
        
        totales.append(total)
    
    # Total general
    total_general = {
        'PAIS': '',
        'NEGOCIO': '',
        'CANAL': '',
        'COBRO/PAGO': '',
        'MONEDA': 'Totales',
        'CLIENTE': ''
    }
    
    for col in columnas_sumar_existentes:
        total_general[col] = df_vencimiento[col].sum()
    
    totales.append(total_general)
    
    # Agregar filas de totales
    df_totales = pd.DataFrame(totales)
    df_vencimiento_con_totales = pd.concat([df_vencimiento, df_totales], ignore_index=True)
    
    print(f"Totales creados: {len(totales)} filas de totalización")
    for total in totales:
        print(f"  - {total['MONEDA']}: {total.get('SALDO TOTAL', 0):,.2f}")
    
    return df_vencimiento_con_totales

# --- Función principal ---
def main():
    # Configurar codificación para stdout y stderr
    import sys
    import codecs
    
    # Configurar stdout para UTF-8
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    else:
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
    
    # Configurar stderr para UTF-8
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')
    else:
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())
    
    print("=== INICIANDO PROCESO MODELO DEUDA ===")
    print(f"Argumentos recibidos: {sys.argv}")
    
    if len(sys.argv) == 5:
        cartera_file = sys.argv[1]
        anticipos_file = sys.argv[2]
        trm_dolar = convertir_valor(sys.argv[3])
        trm_euro = convertir_valor(sys.argv[4])
        carpeta = os.path.dirname(os.path.abspath(__file__))
    else:
        carpeta = os.path.dirname(os.path.abspath(__file__))
        cartera_file = pedir_archivo("Seleccione el archivo de cartera procesada: ", carpeta)
        anticipos_file = pedir_archivo("Seleccione el archivo de anticipos procesados: ", carpeta)
        trm_dolar = convertir_valor(input("Ingrese la TRM Dólar (USD/COP) del último día del mes anterior: "))
        trm_euro = convertir_valor(input("Ingrese la TRM Euro (EUR/COP) del último día del mes anterior: "))

    print(f"Archivo cartera: {cartera_file}")
    print(f"Archivo anticipos: {anticipos_file}")
    print(f"TRM Dólar: {trm_dolar}")
    print(f"TRM Euro: {trm_euro}")
    print(f"Carpeta: {carpeta}")

    # Verificar que los archivos existen
    if not os.path.exists(cartera_file):
        print(f"Error: El archivo de cartera no existe: {cartera_file}")
        return
    if not os.path.exists(anticipos_file):
        print(f"Error: El archivo de anticipos no existe: {anticipos_file}")
        return

    # Leer archivo de cartera (puede ser CSV o Excel)
    try:
        # Intentar leer como Excel primero
        df_cartera = pd.read_excel(cartera_file, engine='openpyxl', dtype=str)
        print("Archivo cartera leído como Excel")
    except Exception as e:
        print(f"Error leyendo como Excel: {e}")
        try:
            # Intentar leer como CSV con coma
            df_cartera = pd.read_csv(cartera_file, sep=',', encoding='latin1', dtype=str)
            print("Archivo cartera leído como CSV con separador ','")
        except Exception as e:
            print(f"Error leyendo como CSV con coma: {e}")
            try:
                # Intentar leer como CSV con punto y coma
                df_cartera = pd.read_csv(cartera_file, sep=';', encoding='latin1', dtype=str)
                print("Archivo cartera leído como CSV con separador ';'")
            except Exception as e:
                print(f"Error leyendo como CSV con punto y coma: {e}")
                # Último intento con detección automática
                try:
                    df_cartera = pd.read_csv(cartera_file, encoding='latin1', dtype=str)
                    print("Archivo cartera leído con detección automática")
                except Exception as e:
                    print(f"Error fatal leyendo archivo de cartera: {e}")
                    return

    # Leer archivo de anticipos
    try:
        df_anticipos = pd.read_excel(anticipos_file, engine='openpyxl')
        print("Archivo anticipos leído correctamente")
    except Exception as e:
        print(f"Error leyendo archivo de anticipos: {e}")
        return

    df_cartera.columns = df_cartera.columns.str.strip()
    df_anticipos.columns = df_anticipos.columns.str.strip()

    # Limpieza y robustez de columnas para cartera y anticipos
    def renombrar_columna(df, nombre_estandar, variantes):
        for col in df.columns:
            # Normalizar nombres de columnas para comparación
            col_normalizada = col.strip().lower().replace(' ', '').replace('_', '').replace('-', '')
            variantes_normalizadas = [v.lower().replace(' ', '').replace('_', '').replace('-', '') for v in variantes]
            
            if col_normalizada in variantes_normalizadas:
                if col != nombre_estandar:
                    print(f"Renombrando columna '{col}' a '{nombre_estandar}'")
                    df.rename(columns={col: nombre_estandar}, inplace=True)
                    return True
        return False

    # --- Columnas clave de cartera ---
    # Columnas clave de cartera con todas las variantes posibles (actualizadas según procesador mejorado)
    columnas_cartera = [
        ('EMPRESA', ['EMPRESA', 'EMPRESA.1']),
        ('ACTIVIDAD', ['ACTIVIDAD']),
        ('CODIGO AGENTE', ['CODIGO AGENTE']),
        ('AGENTE', ['AGENTE']),
        ('CODIGO COBRADOR', ['CODIGO COBRADOR']),
        ('COBRADOR', ['COBRADOR']),
        ('CODIGO CLIENTE', ['CODIGO CLIENTE']),
        ('IDENTIFICACION', ['IDENTIFICACION']),
        ('NOMBRE', ['NOMBRE']),
        ('DENOMINACION COMERCIAL', ['DENOMINACION COMERCIAL', 'DENOMINACIÓN COMERCIAL']),
        ('DIRECCION', ['DIRECCION']),
        ('TELEFONO', ['TELEFONO']),
        ('CIUDAD', ['CIUDAD']),
        ('NUMERO FACTURA', ['NUMERO FACTURA']),
        ('TIPO', ['TIPO']),
        ('FECHA', ['FECHA']),
        ('FECHA VTO', ['FECHA VTO']),
        ('VALOR', ['VALOR']),
        ('SALDO', ['SALDO', ' SALDO ']),
        ('DIA FECHA', ['DIA FECHA']),
        ('MES FECHA', ['MES FECHA']),
        ('AÑO FECHA', ['AÑO FECHA']),
        ('DIA FECHA VTO', ['DIA FECHA VTO']),
        ('MES FECHA VTO', ['MES FECHA VTO']),
        ('AÑO FECHA VTO', ['AÑO FECHA VTO']),
        ('DIAS VENCIDO', ['DIAS VENCIDO']),
        ('DIAS POR VENCER', ['DIAS POR VENCER']),
        ('SALDO VENCIDO', ['SALDO VENCIDO', ' SALDO VENCIDO ']),
        ('% Dotación', ['% Dotación', '%DOTACIÓN']),
        ('  Valor Dotación  ', ['  Valor Dotación  ', 'Valor Dotación']),
        ('Mora Total', ['Mora Total']),
        # Vencimientos históricos (nombres dinámicos según fecha)
        ('jun.-25', ['jun.-25', 'jun-25']),
        ('may.-25', ['may.-25', 'may-25']),
        ('abr.-25', ['abr.-25', 'abr-25']),
        ('mar.-25', ['mar.-25', 'mar-25']),
        ('feb.-25', ['feb.-25', 'feb-25']),
        ('ene.-25', ['ene.-25', 'ene-25']),
        ('Vencimiento_180_dias', ['Vencimiento_180_dias']),
        ('Por_Vencer_1_meses', ['Por_Vencer_1_meses']),
        ('Por_Vencer_2_meses', ['Por_Vencer_2_meses']),
        ('Por_Vencer_3_meses', ['Por_Vencer_3_meses']),
        ('Por_Vencer_+90_dias', ['Por_Vencer_+90_dias']),
        ('Valor Total Por Vencer', ['Valor Total Por Vencer', 'VALOR TOTAL POR VENCER']),
        ('Verificación Suma Saldos', ['Verificación Suma Saldos']),
        ('SALDO NO VENCIDO', ['Saldo no vencido', 'SALDO NO VENCIDO']),
        ('VENCIDO 30', ['VENCIDO 30']),
        ('VENCIDO 60', ['VENCIDO 60']),
        ('VENCIDO 90', ['VENCIDO 90']),
        ('VENCIDO 180', ['VENCIDO 180']),
        ('VENCIDO 360', ['VENCIDO 360']),
        ('VENCIDO + 360', ['VENCIDO + 360']),
        ('Validación Vencimientos', ['Validación Vencimientos']),
        ('  DEUDA INCOBRABLE  ', ['  DEUDA INCOBRABLE  ', 'DEUDA INCOBRABLE']),
        ('MONEDA', ['MONEDA']),
        ('CLIENTE', ['CLIENTE']),
    ]

    # Columnas clave de anticipos (con mapeo de nombres reales)
    columnas_anticipos = [
        ('EMPRESA', ['EMPRESA', 'NCCDEM']),
        ('ACTIVIDAD', ['ACTIVIDAD', 'NCCDAC']),
        ('CODIGO CLIENTE', ['CODIGO CLIENTE', 'NCCDCL']),
        ('NIT/CEDULA', ['NIT/CEDULA', 'WWNIT']),
        ('NOMBRE COMERCIAL', ['NOMBRE COMERCIAL', 'WWNMCL']),
        ('DIRECCION', ['DIRECCION', 'WWNMDO']),
        ('TELEFONO', ['TELEFONO', 'WWTLF1']),
        ('POBLACION', ['POBLACION', 'WWNMPO']),
        ('CODIGO AGENTE', ['CODIGO AGENTE', 'CCCDFB']),
        ('NOMBRE AGENTE', ['NOMBRE AGENTE', 'BDNMNM']),
        ('APELLIDO AGENTE', ['APELLIDO AGENTE', 'BDNMPA']),
        ('TIPO ANTICIPO', ['TIPO ANTICIPO']),
        ('NRO ANTICIPO', ['NRO ANTICIPO', 'NCCDR3']),
        ('VALOR ANTICIPO', ['VALOR ANTICIPO', 'NCIMAN']),
        ('FECHA ANTICIPO', ['FECHA ANTICIPO', 'NCFEGR'])
    ]

    # Función para convertir tipos de datos
    def convertir_tipos_dataframe(df, columnas_numericas):
        for col in columnas_numericas:
            if col in df.columns:
                try:
                    # Convertir a numérico, manteniendo NaN para valores no convertibles
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                    # Llenar NaN con 0
                    df[col] = df[col].fillna(0)
                except Exception as e:
                    print(f"Error convirtiendo columna {col}: {e}")
        return df

    # Renombrar variantes y crear vacías si faltan
    for nombre, variantes in columnas_cartera:
        renombrar_columna(df_cartera, nombre, variantes)
        if nombre not in df_cartera.columns:
            print(f"ADVERTENCIA: No existe la columna '{nombre}' en cartera, se crea vacia.")
            df_cartera[nombre] = ""
    
    # Renombrar variantes y crear vacías si faltan
    for nombre, variantes in columnas_anticipos:
        renombrar_columna(df_anticipos, nombre, variantes)
        if nombre not in df_anticipos.columns:
            print(f"ADVERTENCIA: No existe la columna '{nombre}' en anticipos, se crea vacia.")
            df_anticipos[nombre] = ""

    # Imprimir nombres de columnas para depuración
    print("Columnas disponibles en df_cartera:", df_cartera.columns.tolist())

    # Eliminar columnas duplicadas
    df_cartera = df_cartera.loc[:, ~df_cartera.columns.duplicated()]
    print("Columnas después de eliminar duplicados:", df_cartera.columns.tolist())

    # Si la columna tiene tilde o algún carácter extraño, renombrar
    if 'DENOMINACIÓN COMERCIAL' in df_cartera.columns:
        df_cartera.rename(columns={'DENOMINACIÓN COMERCIAL': 'DENOMINACION COMERCIAL'}, inplace=True)

    # Convertir tipos de datos para columnas numéricas (actualizadas según procesador mejorado)
    columnas_numericas_cartera = [
        'SALDO', 'SALDO VENCIDO', 'SALDO NO VENCIDO', 'VALOR', 'DIAS VENCIDO', 'DIAS POR VENCER',
        'VENCIDO 30', 'VENCIDO 60', 'VENCIDO 90', 'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360',
        '  DEUDA INCOBRABLE  ', '  Valor Dotación  ', 'Mora Total', 'Valor Total Por Vencer',
        'Vencimiento_180_dias', 'Por_Vencer_1_meses', 'Por_Vencer_2_meses', 'Por_Vencer_3_meses',
        'Por_Vencer_+90_dias', 'jun.-25', 'may.-25', 'abr.-25', 'mar.-25', 'feb.-25', 'ene.-25'
    ]
    
    # Validar y corregir valores negativos en cartera
    if 'SALDO' in df_cartera.columns:
        saldos_convertidos = df_cartera['SALDO'].apply(convertir_valor)
        valores_negativos = saldos_convertidos < 0
        if valores_negativos.any():
            print(f"ADVERTENCIA: Se encontraron {valores_negativos.sum()} registros con valores negativos en SALDO de cartera")
            print("Los valores negativos se convertirán a positivos para el procesamiento")
            # Convertir valores negativos a positivos
            df_cartera['SALDO'] = df_cartera['SALDO'].apply(lambda x: str(abs(convertir_valor(x))) if convertir_valor(x) < 0 else x)
    
    columnas_numericas_anticipos = ['VALOR ANTICIPO']
    
    # Validar y corregir valores negativos en anticipos
    if 'VALOR ANTICIPO' in df_anticipos.columns:
        valores_anticipo_convertidos = df_anticipos['VALOR ANTICIPO'].apply(convertir_valor)
        valores_negativos = valores_anticipo_convertidos < 0
        if valores_negativos.any():
            print(f"ADVERTENCIA: Se encontraron {valores_negativos.sum()} registros con valores negativos en VALOR ANTICIPO")
            print("Los valores negativos se convertirán a positivos para el procesamiento")
            # Convertir valores negativos a positivos
            df_anticipos['VALOR ANTICIPO'] = df_anticipos['VALOR ANTICIPO'].apply(lambda x: str(abs(convertir_valor(x))) if convertir_valor(x) < 0 else x)
    
    print("Convirtiendo tipos de datos en cartera...")
    df_cartera = convertir_tipos_dataframe(df_cartera, columnas_numericas_cartera)
    
    print("Convirtiendo tipos de datos en anticipos...")
    df_anticipos = convertir_tipos_dataframe(df_anticipos, columnas_numericas_anticipos)
    
    # Eliminar columnas duplicadas en anticipos también
    df_anticipos = df_anticipos.loc[:, ~df_anticipos.columns.duplicated()]
    print("Columnas de anticipos después de eliminar duplicados:", df_anticipos.columns.tolist())

    # Validar datos según reglas del modelo de deuda
    print("\n=== VALIDANDO DATOS ===")
    datos_validos = validar_datos_modelo_deuda(df_cartera, df_anticipos)
    if not datos_validos:
        print("ADVERTENCIA: Se encontraron errores en los datos, pero se continuara con el procesamiento")

    # Asegurar que la columna MONEDA exista
    if 'MONEDA' not in df_cartera.columns:
        for col in df_cartera.columns:
            if 'moneda' in col.lower() and col != 'MONEDA':
                print(f"Renombrando columna '{col}' a 'MONEDA'")
                df_cartera.rename(columns={col: 'MONEDA'}, inplace=True)
        if 'MONEDA' not in df_cartera.columns:
            print("ADVERTENCIA: No existe la columna 'MONEDA', se crea vacia.")
            df_cartera['MONEDA'] = ""

    # --- Filtrar líneas según reglas exactas ---
    # Líneas en pesos (según reglas)
    lineas_pesos = ['80', '41', '44', '47', '10', '15', '20', '21', '23', '25', '28', '29', '31', '32', '53', '56', '60', '62', '63', '64', '65', '66', '69']
    
    # Líneas en divisas (según reglas)
    lineas_divisas = ['11', '18', '57', '41']  # Dólares y Euros
    
    # Filtrar cartera para pesos
    df_pesos = df_cartera[df_cartera['ACTIVIDAD'].astype(str).isin(lineas_pesos)].copy()
    
    # Filtrar cartera para divisas
    df_divisas = df_cartera[df_cartera['ACTIVIDAD'].astype(str).isin(lineas_divisas)].copy()
    
    # Aplicar TRM a divisas según especificaciones exactas
    if 'MONEDA' in df_divisas.columns:
        columnas_aplicar_trm = ['SALDO', 'SALDO VENCIDO', 'SALDO NO VENCIDO', 'VALOR', 
                               'VENCIDO 30', 'VENCIDO 60', 'VENCIDO 90', 'VENCIDO 180', 
                               'VENCIDO 360', 'VENCIDO + 360', '  DEUDA INCOBRABLE  ', 
                               '  Valor Dotación  ', 'Mora Total', 'Valor Total Por Vencer']
        
        # Aplicar TRM según la moneda (Dólares y Euros)
        for moneda in ['DÓLAR', 'DOLAR', 'USD']:
            mask = df_divisas['MONEDA'].str.upper().isin([moneda, 'DOLAR', 'USD'])
            if mask.any():
                for col in columnas_aplicar_trm:
                    if col in df_divisas.columns:
                        try:
                            # Aplicar TRM Dólar solo a valores numéricos válidos
                            df_divisas.loc[mask, col] = df_divisas.loc[mask, col] * trm_dolar
                            print(f"Aplicando TRM Dólar ({trm_dolar}) a columna '{col}' para moneda '{moneda}'")
                        except Exception as e:
                            print(f"Error aplicando TRM Dólar a columna {col}: {e}")
        
        for moneda in ['EURO', 'EUR']:
            mask = df_divisas['MONEDA'].str.upper().isin([moneda, 'EURO', 'EUR'])
            if mask.any():
                for col in columnas_aplicar_trm:
                    if col in df_divisas.columns:
                        try:
                            # Aplicar TRM Euro solo a valores numéricos válidos
                            df_divisas.loc[mask, col] = df_divisas.loc[mask, col] * trm_euro
                            print(f"Aplicando TRM Euro ({trm_euro}) a columna '{col}' para moneda '{moneda}'")
                        except Exception as e:
                            print(f"Error aplicando TRM Euro a columna {col}: {e}")
        
        # Aplicar TRM a anticipos en divisas
        mask_anticipos = df_divisas['MONEDA'].str.upper().isin(['DOLAR', 'DÓLAR', 'USD'])
        if mask_anticipos.any():
            for col in columnas_aplicar_trm:
                if col in df_divisas.columns:
                    try:
                        # Los anticipos en divisas se convierten con TRM Dólar
                        df_divisas.loc[mask_anticipos, col] = df_divisas.loc[mask_anticipos, col] * trm_dolar
                        print(f"Aplicando TRM Dólar ({trm_dolar}) a anticipos en divisas")
                    except Exception as e:
                        print(f"Error aplicando TRM Dólar a anticipos: {e}")

    # --- Crear hoja de vencimiento ---
    print("Procesando hoja VENCIMIENTO...")
    df_vencimiento = crear_hoja_vencimiento(df_cartera)
    df_vencimiento = crear_totales_por_moneda(df_vencimiento)
    print(f"Registros en hoja VENCIMIENTO: {len(df_vencimiento)}")

    # --- Asegura estructura y columnas correctas por hoja ---
    # Columnas para pesos y divisas (actualizadas según procesador mejorado)
    columnas_operacion = [
        'EMPRESA', 'ACTIVIDAD', 'DENOMINACION COMERCIAL', 'SALDO', 'SALDO VENCIDO', '% Dotación',
        'Valor Total Por Vencer', 'SALDO NO VENCIDO', 'VENCIDO 30', 'VENCIDO 60', 'VENCIDO 90',
        'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360', '  DEUDA INCOBRABLE  ', 'MONEDA'
    ]

    # --- Anticipos para pesos según especificaciones exactas ---
    anticipos_pesos = pd.DataFrame(columns=columnas_operacion)
    for _, row in df_anticipos.iterrows():
        nuevo = {col: 0 for col in columnas_operacion}
        nuevo['EMPRESA'] = str(row.get('EMPRESA', ''))
        nuevo['ACTIVIDAD'] = str(row.get('ACTIVIDAD', ''))
        nuevo['DENOMINACION COMERCIAL'] = str(row.get('NOMBRE COMERCIAL', ''))
        
        # Convertir valor de anticipo usando la función convertir_valor
        valor_anticipo = row.get('VALOR ANTICIPO', 0)
        if isinstance(valor_anticipo, str):
            valor_anticipo = convertir_valor(valor_anticipo)
        else:
            valor_anticipo = float(valor_anticipo)
        
        nuevo['SALDO'] = valor_anticipo
        nuevo['Valor Total Por Vencer'] = valor_anticipo
        nuevo['SALDO NO VENCIDO'] = valor_anticipo
        nuevo['MONEDA'] = 'PESOS COL'
        
        nuevo_df = pd.DataFrame([nuevo])
        # Asegurar que tenga las mismas columnas
        for col in columnas_operacion:
            if col not in nuevo_df.columns:
                nuevo_df[col] = 0
        anticipos_pesos = pd.concat([anticipos_pesos, nuevo_df], ignore_index=True)
    
    if not anticipos_pesos.empty:
        df_pesos = pd.concat([df_pesos, anticipos_pesos], ignore_index=True)
        print(f"Agregados {len(anticipos_pesos)} registros de anticipos a pesos")

    # --- Anticipos para divisas según especificaciones exactas ---
    # Los anticipos también van a divisas pero convertidos con TRM
    anticipos_divisas = pd.DataFrame(columns=columnas_operacion)
    for _, row in df_anticipos.iterrows():
        nuevo = {col: 0 for col in columnas_operacion}
        nuevo['EMPRESA'] = str(row.get('EMPRESA', ''))
        nuevo['ACTIVIDAD'] = str(row.get('ACTIVIDAD', ''))
        nuevo['DENOMINACION COMERCIAL'] = str(row.get('NOMBRE COMERCIAL', ''))
        
        # Convertir valor de anticipo usando la función convertir_valor
        valor_anticipo = row.get('VALOR ANTICIPO', 0)
        if isinstance(valor_anticipo, str):
            valor_anticipo = convertir_valor(valor_anticipo)
        else:
            valor_anticipo = float(valor_anticipo)
        
        nuevo['SALDO'] = valor_anticipo
        nuevo['Valor Total Por Vencer'] = valor_anticipo
        nuevo['SALDO NO VENCIDO'] = valor_anticipo
        nuevo['MONEDA'] = 'DOLAR'  # Los anticipos se convierten a dólares
        
        nuevo_df = pd.DataFrame([nuevo])
        # Asegurar que tenga las mismas columnas
        for col in columnas_operacion:
            if col not in nuevo_df.columns:
                nuevo_df[col] = 0
        anticipos_divisas = pd.concat([anticipos_divisas, nuevo_df], ignore_index=True)
    
    if not anticipos_divisas.empty:
        df_divisas = pd.concat([df_divisas, anticipos_divisas], ignore_index=True)
        print(f"Agregados {len(anticipos_divisas)} registros de anticipos a divisas")

    for col in columnas_operacion:
        if col not in df_pesos.columns:
            df_pesos[col] = ""
        if col not in df_divisas.columns:
            df_divisas[col] = ""
    df_pesos = df_pesos[columnas_operacion]
    df_divisas = df_divisas[columnas_operacion]

    # Columnas para vencimiento según el cuadro especificado (actualizadas según procesador mejorado)
    columnas_vencimiento = [
        'PAIS', 'NEGOCIO', 'CANAL', 'COBRO/PAGO', 'MONEDA', 'CLIENTE',
        'SALDO TOTAL', 'SALDO NO VENCIDO', 'VENCIDO 30', 'VENCIDO 60', 'VENCIDO 90',
        'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360', '  DEUDA INCOBRABLE  '
    ]
    for col in columnas_vencimiento:
        if col not in df_vencimiento.columns:
            df_vencimiento[col] = ""
    df_vencimiento = df_vencimiento[columnas_vencimiento]

    # Llena vacíos en columnas numéricas con 0 (actualizadas según procesador mejorado)
    columnas_numericas = ['SALDO', 'SALDO VENCIDO', 'Valor Total Por Vencer', 'SALDO NO VENCIDO', 
                          'VENCIDO 30', 'VENCIDO 60', 'VENCIDO 90', 'VENCIDO 180', 'VENCIDO 360', 
                          'VENCIDO + 360', '  DEUDA INCOBRABLE  ', '% Dotación', '  Valor Dotación  ', 'Mora Total']
    
    for df in [df_pesos, df_divisas]:
        for col in columnas_numericas:
            if col in df.columns and df[col].dtype == object:
                try:
                    df[col] = df[col].apply(convertir_valor)
                except Exception:
                    pass

    # Aplicar formato colombiano a los números
    columnas_numericas_pesos = ['SALDO', 'SALDO VENCIDO', 'Valor Total Por Vencer', 'SALDO NO VENCIDO', 
                                'VENCIDO 30', 'VENCIDO 60', 'VENCIDO 90', 'VENCIDO 180', 'VENCIDO 360', 
                                'VENCIDO + 360', '  DEUDA INCOBRABLE  ']
    
    columnas_numericas_vencimiento = ['SALDO TOTAL', 'SALDO NO VENCIDO', 'VENCIDO 30', 'VENCIDO 60', 
                                      'VENCIDO 90', 'VENCIDO 180', 'VENCIDO 360', 'VENCIDO + 360', 
                                      '  DEUDA INCOBRABLE  ']
    
    # Columnas de porcentaje
    columnas_porcentaje = ['% Dotación']
    
    # Aplicar formato a pesos y divisas (actualizadas según procesador mejorado)
    df_pesos = aplicar_formato_colombiano_dataframe(df_pesos, columnas_numericas_pesos)
    df_divisas = aplicar_formato_colombiano_dataframe(df_divisas, columnas_numericas_pesos)
    
    # Aplicar formato a vencimiento
    df_vencimiento = aplicar_formato_colombiano_dataframe(df_vencimiento, columnas_numericas_vencimiento)
    
    # Aplicar formato especial a porcentajes
    for df in [df_pesos, df_divisas]:
        for col in columnas_porcentaje:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: formatear_numero_colombiano(x, es_porcentaje=True))

    print(f"Registros en hoja PESOS: {len(df_pesos)}")
    print(f"Registros en hoja DIVISAS: {len(df_divisas)}")
    print(f"Registros en hoja VENCIMIENTO: {len(df_vencimiento)}")

    # --- Guardar Excel con tres hojas ---
    carpeta = os.path.join(os.path.dirname(__file__), '..', 'PROVCA_PROCESADOS')
    os.makedirs(carpeta, exist_ok=True)
    
    # Generar nombre exacto según especificaciones: "1 Modelo Deuda"
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    ruta_salida = os.path.join(carpeta, f'1_Modelo_Deuda_{timestamp}.xlsx')
    
    # Verificar que los DataFrames no estén vacíos
    if df_pesos.empty and df_divisas.empty and df_vencimiento.empty:
        print("ERROR: Todos los DataFrames están vacíos. No se puede generar archivo.")
        return
    
    print(f"DataFrames validados:")
    print(f"- PESOS: {len(df_pesos)} filas, {len(df_pesos.columns)} columnas")
    print(f"- DIVISAS: {len(df_divisas)} filas, {len(df_divisas.columns)} columnas")
    print(f"- VENCIMIENTO: {len(df_vencimiento)} filas, {len(df_vencimiento.columns)} columnas")
    
    # Guardar con la función de múltiples hojas
    if guardar_excel_con_hojas(df_pesos, df_divisas, df_vencimiento, ruta_salida):
        print(f"Archivo generado: {ruta_salida}")
        print("=== PROCESO COMPLETADO EXITOSAMENTE ===")
        print(f"Hoja PESOS: {len(df_pesos)} registros")
        print(f"Hoja DIVISAS: {len(df_divisas)} registros")
        print(f"Hoja VENCIMIENTO: {len(df_vencimiento)} registros")
    else:
        print("Error: No se pudo generar el archivo correctamente")
        return
    
    # Limpiar memoria
    del df_pesos, df_divisas, df_vencimiento, df_cartera, df_anticipos
    gc.collect()

if __name__ == "__main__":
    main() 