#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Modelo de Deuda - Versión Corregida Final (anti-styles.xml)
Genera archivo Excel con 3 hojas: PESOS, DIVISAS y VENCIMIENTO
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# =======================
# Utilidades de estilos
# =======================

def ensure_named_style(workbook, style: NamedStyle):
    """Registra un NamedStyle si no existe aún en el workbook (evita duplicados)."""
    existing = {s if isinstance(s, str) else s.name for s in workbook.named_styles}
    if style.name not in existing:
        workbook.add_named_style(style)

def build_styles(workbook):
    """
    Crea/asegura estilos reutilizables:
      - hdr: encabezado fondo azul, texto blanco, centrado
      - num: número con separador de miles y 2 decimales, alineado a la derecha
      - num_int: número entero con miles (no lo usamos por defecto, pero queda disponible)
      - total_row: totales por moneda (fondo azul claro + negrita)
      - grand_total_row: total general (fondo verde claro + negrita)
    """
    hdr = NamedStyle(name="hdr")
    hdr.font = Font(bold=True, color="FFFFFF")
    hdr.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    hdr.alignment = Alignment(horizontal="center", vertical="center")

    num = NamedStyle(name="num")
    num.number_format = '#,##0.00_);[Red](#,##0.00);-_'
    num.alignment = Alignment(horizontal='right', vertical='center')

    num_int = NamedStyle(name="num_int")
    num_int.number_format = '#,##0_);[Red](#,##0);-_'
    num_int.alignment = Alignment(horizontal='right', vertical='center')

    total_row = NamedStyle(name="total_row")
    total_row.font = Font(bold=True)
    total_row.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

    grand_total_row = NamedStyle(name="grand_total_row")
    grand_total_row.font = Font(bold=True)
    grand_total_row.fill = PatternFill(start_color="D4E6B7", end_color="D4E6B7", fill_type="solid")

    for st in (hdr, num, num_int, total_row, grand_total_row):
        ensure_named_style(workbook, st)

# =======================
# PROCESAMIENTO PRINCIPAL
# =======================

def procesar_modelo_deuda(cartera_path, anticipos_path, trm_dolar, trm_euro, output_path):
    """
    Procesa los archivos de cartera y anticipos para generar el formato de deuda consolidado
    """
    try:
        print(f"📄 Iniciando procesamiento del modelo de deuda...")
        print(f"📁 Cartera: {cartera_path}")
        print(f"📁 Anticipos: {anticipos_path}")
        print(f"💱 TRM USD: {trm_dolar:,.2f}")
        print(f"💱 TRM EUR: {trm_euro:,.2f}")
        
        # Leer archivos
        df_cartera = leer_archivo(cartera_path, "Cartera")
        df_anticipos = leer_archivo(anticipos_path, "Anticipos")
        
        print(f"📊 Cartera - Registros: {len(df_cartera)}, Columnas: {len(df_cartera.columns)}")
        print(f"📊 Anticipos - Registros: {len(df_anticipos)}, Columnas: {len(df_anticipos.columns)}")
        
        # Normalizar columnas
        df_cartera = normalizar_columnas_cartera(df_cartera)
        df_anticipos = normalizar_columnas_anticipos(df_anticipos)
        
        # Crear campo LINEA_VENTA correctamente
        df_cartera = crear_linea_venta(df_cartera)
        
        # Debug: mostrar líneas encontradas
        print(f"🔍 Líneas únicas en cartera:")
        lineas_unicas = df_cartera['LINEA_VENTA'].unique()
        for linea in sorted(lineas_unicas):
            count = len(df_cartera[df_cartera['LINEA_VENTA'] == linea])
            print(f"   {linea}: {count} registros")
        
        # Definir líneas según especificaciones
        lineas_pesos = [
            'CT80', 'ED41', 'ED44', 'ED47', 'PL10', 'PL15', 'PL20', 'PL21', 'PL23', 'PL25',
            'PL28', 'PL29', 'PL31', 'PL32', 'PL53', 'PL56', 'PL60', 'PL62', 'PL63', 'PL64',
            'PL65', 'PL66', 'PL69'
        ]
        
        lineas_divisas = ['PL11', 'PL18', 'PL57', 'PL41']
        
        # Filtrar cartera
        df_cartera_pesos = df_cartera[df_cartera['LINEA_VENTA'].isin(lineas_pesos)].copy()
        df_cartera_divisas = df_cartera[df_cartera['LINEA_VENTA'].isin(lineas_divisas)].copy()
        
        print(f"📊 Cartera PESOS: {len(df_cartera_pesos)} registros")
        print(f"📊 Cartera DIVISAS: {len(df_cartera_divisas)} registros")
        
        if len(df_cartera_divisas) > 0:
            print("🔍 Líneas encontradas en DIVISAS:")
            for linea in df_cartera_divisas['LINEA_VENTA'].unique():
                print(f"   {linea}")
        
        # Procesar anticipos
        df_anticipos_procesado = procesar_anticipos_para_cartera(df_anticipos)
        
        # Crear hojas
        df_pesos = crear_hoja_pesos(df_cartera_pesos, df_anticipos_procesado)
        df_divisas = crear_hoja_divisas(df_cartera_divisas, df_anticipos_procesado, trm_dolar, trm_euro)
        df_vencimiento = crear_hoja_vencimiento(df_cartera_pesos, df_cartera_divisas, df_anticipos_procesado, trm_dolar, trm_euro)
        
        # Crear archivo Excel
        crear_excel_final(df_pesos, df_divisas, df_vencimiento, output_path)
        
        print(f"✅ Archivo generado exitosamente: {output_path}")
        return os.path.abspath(output_path)
        
    except Exception as e:
        print(f"❌ Error en procesamiento: {e}")
        raise ValueError(f"Error al procesar modelo de deuda: {e}")

# =======================
# LECTURA
# =======================

def leer_archivo(file_path, tipo_archivo):
    """Lee archivo CSV o Excel de manera robusta"""
    try:
        if str(file_path).lower().endswith('.csv'):
            for sep in [';', ',', '\t']:
                for enc in ['latin1', 'utf-8', 'cp1252']:
                    try:
                        df = pd.read_csv(file_path, sep=sep, encoding=enc, dtype=str)
                        if len(df.columns) > 1:
                            print(f"✅ {tipo_archivo} CSV leído: sep='{sep}', encoding='{enc}'")
                            return df
                    except:
                        continue
            raise ValueError(f"No se pudo leer el archivo CSV {tipo_archivo}")
        else:
            df = pd.read_excel(file_path, engine='openpyxl', dtype=str)
            print(f"✅ {tipo_archivo} Excel leído correctamente")
            return df
    except Exception as e:
        raise ValueError(f"Error leyendo {tipo_archivo} {file_path}: {e}")

# =======================
# NORMALIZACIÓN
# =======================

def normalizar_columnas_cartera(df):
    """Normaliza columnas del archivo de cartera"""
    print("🔧 Normalizando columnas de cartera...")
    
    # Mapeo de columnas comunes
    column_mapping = {
        'PCCDEM': 'EMPRESA',
        'PCCDAC': 'ACTIVIDAD',
        'PCDEAC': 'DESC_EMPRESA',
        'PCCDCL': 'CODIGO_CLIENTE',
        'PCNMCM': 'DENOMINACION_COMERCIAL',
        'PCSALD': 'SALDO_TOTAL',
        'CODIGO CLIENTE': 'CODIGO_CLIENTE',
        'NOMBRE COMERCIAL': 'NOMBRE_COMERCIAL', 
        'DENOMINACION COMERCIAL': 'DENOMINACION_COMERCIAL',
        'SALDO TOTAL': 'SALDO_TOTAL',
        'SALDO': 'SALDO_TOTAL',
        'SALDO NO VENCIDO': 'SALDO_NO_VENCIDO',
        'VENCIDO 30': 'VENCIDO_30',
        'VENCIDO 60': 'VENCIDO_60',
        'VENCIDO 90': 'VENCIDO_90', 
        'VENCIDO 180': 'VENCIDO_180',
        'VENCIDO 360': 'VENCIDO_360',
        'VENCIDO + 360': 'VENCIDO_360_PLUS',
        'DEUDA INCOBRABLE': 'DEUDA_INCOBRABLE'
    }
    
    df = df.rename(columns=column_mapping)
    
    # Convertir columnas numéricas
    columnas_numericas = [
        'SALDO_TOTAL', 'SALDO_NO_VENCIDO', 'VENCIDO_30', 'VENCIDO_60', 'VENCIDO_90',
        'VENCIDO_180', 'VENCIDO_360', 'VENCIDO_360_PLUS', 'DEUDA_INCOBRABLE'
    ]
    
    for col in columnas_numericas:
        if col in df.columns:
            df[col] = convertir_a_numerico(df[col])
        else:
            df[col] = 0.0
    
    # Campos obligatorios con valores por defecto
    if 'EMPRESA' not in df.columns:
        df['EMPRESA'] = 'PL'
    if 'ACTIVIDAD' not in df.columns:
        df['ACTIVIDAD'] = '99'
    if 'CODIGO_CLIENTE' not in df.columns:
        df['CODIGO_CLIENTE'] = 'SIN_CODIGO'
    
    # Calcular SALDO_NO_VENCIDO si no existe o está en cero
    if 'SALDO_NO_VENCIDO' not in df.columns or df['SALDO_NO_VENCIDO'].sum() == 0:
        vencidos = df[['VENCIDO_30', 'VENCIDO_60', 'VENCIDO_90', 'VENCIDO_180', 'VENCIDO_360', 'VENCIDO_360_PLUS']].sum(axis=1)
        df['SALDO_NO_VENCIDO'] = df['SALDO_TOTAL'] - vencidos
        df['SALDO_NO_VENCIDO'] = df['SALDO_NO_VENCIDO'].clip(lower=0)
    
    print(f"✅ Cartera normalizada: {len(df)} registros")
    return df

def crear_linea_venta(df):
    """Crea la línea de venta combinando EMPRESA + ACTIVIDAD"""
    df['EMPRESA'] = df['EMPRESA'].astype(str).str.strip()
    df['ACTIVIDAD'] = df['ACTIVIDAD'].astype(str).str.strip().str.zfill(2)
    df['LINEA_VENTA'] = df['EMPRESA'] + df['ACTIVIDAD']
    return df

def normalizar_columnas_anticipos(df):
    """Normaliza columnas del archivo de anticipos"""
    print("🔧 Normalizando columnas de anticipos...")
    
    column_mapping = {
        'NCCDEM': 'EMPRESA',
        'NCCDAC': 'ACTIVIDAD', 
        'NCCDCL': 'CODIGO_CLIENTE',
        'NCIMAN': 'VALOR_ANTICIPO',
        'VALOR_ANTICIPO': 'VALOR_ANTICIPO',
        'VALOR ANTICIPO': 'VALOR_ANTICIPO',
        'CODIGO CLIENTE': 'CODIGO_CLIENTE'
    }
    
    df = df.rename(columns=column_mapping)
    
    if 'VALOR_ANTICIPO' in df.columns:
        df['VALOR_ANTICIPO'] = convertir_a_numerico(df['VALOR_ANTICIPO'])
        # Los anticipos son negativos (dinero adelantado por el cliente)
        df['VALOR_ANTICIPO'] = df['VALOR_ANTICIPO'].abs() * -1
    else:
        df['VALOR_ANTICIPO'] = 0.0
    
    # Campos por defecto
    if 'EMPRESA' not in df.columns:
        df['EMPRESA'] = 'AN'  # Anticipos
    if 'ACTIVIDAD' not in df.columns:
        df['ACTIVIDAD'] = '00'
    if 'CODIGO_CLIENTE' not in df.columns:
        df['CODIGO_CLIENTE'] = 'ANTICIPO_SIN_CODIGO'
    
    print(f"✅ Anticipos normalizados: {len(df)} registros")
    return df

def convertir_a_numerico(serie):
    """Convierte serie a valores numéricos manejando formato colombiano"""
    def convertir_valor(x):
        if pd.isna(x) or x == '':
            return 0.0
        s = str(x).strip().replace('$', '').replace(' ', '').replace('\u200b', '')
        if s == '' or s == '-':
            return 0.0
        # Manejar formato colombiano: 1.234.567,89
        if '.' in s and ',' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s and '.' not in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except:
            return 0.0
    return serie.apply(convertir_valor)

def procesar_anticipos_para_cartera(df_anticipos):
    """Convierte anticipos al formato de cartera"""
    df = df_anticipos.copy()
    if 'VALOR_ANTICIPO' in df.columns:
        df['SALDO_TOTAL'] = df['VALOR_ANTICIPO']
    else:
        df['SALDO_TOTAL'] = 0.0
    # Para anticipos, todo está "no vencido" (son negativos)
    df['SALDO_NO_VENCIDO'] = df['SALDO_TOTAL']
    # Todos los vencimientos en cero para anticipos
    for col in ['VENCIDO_30', 'VENCIDO_60', 'VENCIDO_90', 'VENCIDO_180', 'VENCIDO_360', 'VENCIDO_360_PLUS', 'DEUDA_INCOBRABLE']:
        df[col] = 0.0
    # Crear LINEA_VENTA para anticipos
    df = crear_linea_venta(df)
    return df

# =======================
# CREACIÓN DE HOJAS
# =======================

def crear_hoja_pesos(df_cartera_pesos, df_anticipos):
    """Crea la hoja PESOS consolidando cartera y anticipos"""
    print("📄 Creando hoja PESOS...")
    # Consolidar solo anticipos que no sean divisas
    df_anticipos_pesos = df_anticipos[~df_anticipos['LINEA_VENTA'].isin(['PL11', 'PL18', 'PL57', 'PL41'])]
    df_pesos = pd.concat([df_cartera_pesos, df_anticipos_pesos], ignore_index=True)
    # Eliminar registros con saldo cero
    df_pesos = df_pesos[df_pesos['SALDO_TOTAL'].abs() > 0.01]
    print(f"✅ Hoja PESOS creada: {len(df_pesos)} registros")
    return df_pesos

def crear_hoja_divisas(df_cartera_divisas, df_anticipos, trm_dolar, trm_euro):
    """Crea la hoja DIVISAS con conversión a pesos"""
    print("📄 Creando hoja DIVISAS...")
    # Incluir anticipos que correspondan a divisas
    df_anticipos_divisas = df_anticipos[df_anticipos['LINEA_VENTA'].isin(['PL11', 'PL18', 'PL57', 'PL41'])]
    # Consolidar
    df_divisas = pd.concat([df_cartera_divisas, df_anticipos_divisas], ignore_index=True)
    if df_divisas.empty:
        print("⚠️ No hay registros de divisas")
        # Crear DataFrame vacío con estructura correcta
        df_divisas = pd.DataFrame(columns=[
            'EMPRESA', 'ACTIVIDAD', 'LINEA_VENTA', 'CODIGO_CLIENTE', 'SALDO_TOTAL',
            'SALDO_NO_VENCIDO', 'VENCIDO_30', 'VENCIDO_60', 'VENCIDO_90',
            'VENCIDO_180', 'VENCIDO_360', 'VENCIDO_360_PLUS', 'DEUDA_INCOBRABLE'
        ])
    else:
        # Aplicar conversión TRM según línea de venta
        columnas_convertir = ['SALDO_TOTAL', 'SALDO_NO_VENCIDO', 'VENCIDO_30', 'VENCIDO_60', 
                              'VENCIDO_90', 'VENCIDO_180', 'VENCIDO_360', 'VENCIDO_360_PLUS', 'DEUDA_INCOBRABLE']
        for col in columnas_convertir:
            df_divisas[col] = df_divisas.apply(
                lambda row: (row[col] * trm_euro if row['LINEA_VENTA'] == 'PL41' else row[col] * trm_dolar), 
                axis=1
            )
        # Eliminar registros con saldo cero
        df_divisas = df_divisas[df_divisas['SALDO_TOTAL'].abs() > 0.01]
    print(f"✅ Hoja DIVISAS creada: {len(df_divisas)} registros")
    return df_divisas

def crear_hoja_vencimiento(df_cartera_pesos, df_cartera_divisas, df_anticipos, trm_dolar, trm_euro):
    """Crea la hoja VENCIMIENTO consolidada"""
    print("📄 Creando hoja VENCIMIENTO...")
    # Combinar todos los datos
    df_consolidado = pd.concat([df_cartera_pesos, df_cartera_divisas, df_anticipos], ignore_index=True)
    # Mapeo negocio-canal según documento
    negocio_canal_mapping = {
        'PL15': 'E-COMMERCE',
        'PL20': 'LIBRERIAS 1', 'PL25': 'LIBRERIAS 1',
        'PL10': 'LIBRERIAS 2', 'PL21': 'LIBRERIAS 2', 
        'PL53': 'LIBRERIAS 3', 'PL63': 'LIBRERIAS 3',
        'PL66': 'OTOS DIGITAL',
        'PL60': 'OTROS', 'PL64': 'OTROS', 'PL65': 'OTROS',
        'PL28': 'SALDOS', 'PL29': 'SALDOS', 'PL31': 'SALDOS',
        'PL18': 'EXPORTACION', 'PL11': 'EXPORTACION',
        'PL57': 'AULA',
        'PL41': 'OTR', 
        'CT80': 'TINTA CLUB DEL LIBRO',
        'AN00': 'ANTICIPOS'  # Para anticipos
    }
    df_consolidado['NEGOCIO'] = df_consolidado['LINEA_VENTA'].map(negocio_canal_mapping).fillna('OTROS')
    df_consolidado['CANAL'] = df_consolidado['NEGOCIO']
    df_consolidado['Pais'] = 'Colombia'
    df_consolidado['COBRO/PAGO'] = 'CLIENTE'
    # Determinar moneda según línea
    def determinar_moneda(linea_venta):
        if linea_venta in ['PL11', 'PL18', 'PL57']:
            return 'DOLAR'
        elif linea_venta == 'PL41':
            return 'EURO'
        else:
            return 'PESOS COL'
    df_consolidado['MONEDA'] = df_consolidado['LINEA_VENTA'].apply(determinar_moneda)
    # Convertir divisas a pesos para consolidación
    df_consolidado_pesos = df_consolidado.copy()
    mask_dolar = df_consolidado_pesos['MONEDA'] == 'DOLAR'
    mask_euro = df_consolidado_pesos['MONEDA'] == 'EURO'
    columnas_convertir = ['SALDO_TOTAL', 'SALDO_NO_VENCIDO', 'VENCIDO_30', 'VENCIDO_60',
                          'VENCIDO_90', 'VENCIDO_180', 'VENCIDO_360', 'VENCIDO_360_PLUS', 'DEUDA_INCOBRABLE']
    for col in columnas_convertir:
        df_consolidado_pesos.loc[mask_dolar, col] *= trm_dolar
        df_consolidado_pesos.loc[mask_euro, col] *= trm_euro
    # Agrupar por dimensiones (usando datos convertidos a pesos)
    cols_agrupacion = ['Pais', 'NEGOCIO', 'CANAL', 'COBRO/PAGO', 'MONEDA', 'CODIGO_CLIENTE']
    agg_dict = {col: 'sum' for col in columnas_convertir}
    df_agrupado = df_consolidado_pesos.groupby(cols_agrupacion).agg(agg_dict).reset_index()
    # Renombrar columnas para salida
    df_agrupado = df_agrupado.rename(columns={
        'CODIGO_CLIENTE': 'CLIENTE',
        'SALDO_TOTAL': 'SALDO TOTAL',
        'SALDO_NO_VENCIDO': 'Saldo No vencido',
        'VENCIDO_30': 'Vencido 30',
        'VENCIDO_60': 'Vencido 60',
        'VENCIDO_90': 'Vencido 90',
        'VENCIDO_180': 'Vencido 180',
        'VENCIDO_360': 'Vencido 360',
        'VENCIDO_360_PLUS': 'Vencido + 360',
        'DEUDA_INCOBRABLE': 'DEUDA INCOBRABLE'
    })
    # Crear totales por moneda
    df_totales = crear_totales_por_moneda(df_agrupado, trm_dolar, trm_euro)
    df_final = pd.concat([df_agrupado, df_totales], ignore_index=True)
    print(f"✅ Hoja VENCIMIENTO creada: {len(df_final)} registros")
    return df_final

def crear_totales_por_moneda(df_agrupado, trm_dolar, trm_euro):
    """Crea totales por moneda"""
    print("📊 Creando totales por moneda...")
    totales = []
    cols_suma = ['SALDO TOTAL', 'Saldo No vencido', 'Vencido 30', 'Vencido 60',
                 'Vencido 90', 'Vencido 180', 'Vencido 360', 'Vencido + 360', 'DEUDA INCOBRABLE']
    # Total Pesos (ya están en pesos)
    df_pesos = df_agrupado[df_agrupado['MONEDA'] == 'PESOS COL']
    if not df_pesos.empty:
        total_pesos = {col: df_pesos[col].sum() for col in cols_suma}
        totales.append({
            'Pais': '', 'NEGOCIO': '', 'CANAL': '', 'COBRO/PAGO': '',
            'MONEDA': 'Moneda Local', 'CLIENTE': '',
            **total_pesos
        })
    # Total Dólar (ya convertidos a pesos)
    df_dolar = df_agrupado[df_agrupado['MONEDA'] == 'DOLAR'] 
    if not df_dolar.empty:
        total_dolar_pesos = {col: df_dolar[col].sum() for col in cols_suma}
        valor_original_usd = total_dolar_pesos['SALDO TOTAL'] / trm_dolar
        totales.append({
            'Pais': '', 'NEGOCIO': '', 'CANAL': '', 'COBRO/PAGO': '',
            'MONEDA': 'Dólar', 'CLIENTE': f"{valor_original_usd:,.2f}",
            **total_dolar_pesos
        })
    # Total Euro (ya convertidos a pesos)
    df_euro = df_agrupado[df_agrupado['MONEDA'] == 'EURO']
    if not df_euro.empty:
        total_euro_pesos = {col: df_euro[col].sum() for col in cols_suma}
        valor_original_eur = total_euro_pesos['SALDO TOTAL'] / trm_euro
        totales.append({
            'Pais': '', 'NEGOCIO': '', 'CANAL': '', 'COBRO/PAGO': '',
            'MONEDA': 'Euro', 'CLIENTE': f"{valor_original_eur:,.2f}",
            **total_euro_pesos
        })
    # Total general
    if totales:
        total_general = {col: sum(t.get(col, 0) for t in totales if isinstance(t.get(col), (int, float, float))) 
                         for col in cols_suma}
        totales.append({
            'Pais': '', 'NEGOCIO': '', 'CANAL': '', 'COBRO/PAGO': '',
            'MONEDA': '', 'CLIENTE': 'Totales',
            **total_general
        })
    return pd.DataFrame(totales)

# =======================
# CREACIÓN DE EXCEL
# =======================

def crear_excel_final(df_pesos, df_divisas, df_vencimiento, output_path):
    """Crea el archivo Excel final con las 3 hojas y aplica formatos sin explotar styles.xml"""
    print("📝 Creando archivo Excel...")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Escribir las 3 hojas
        df_pesos.to_excel(writer, sheet_name='PESOS', index=False)
        df_divisas.to_excel(writer, sheet_name='DIVISAS', index=False)
        df_vencimiento.to_excel(writer, sheet_name='VENCIMIENTO', index=False)
        
        # Estilos reutilizables
        workbook = writer.book
        build_styles(workbook)

        # Aplicar formato
        ws_p = workbook['PESOS']
        ws_d = workbook['DIVISAS']
        ws_v = workbook['VENCIMIENTO']

        aplicar_formato_hoja(ws_p)
        aplicar_formato_hoja(ws_d) 
        aplicar_formato_hoja_vencimiento(ws_v)

def aplicar_formato_hoja(worksheet):
    """Aplica formato básico a las hojas usando NamedStyle reutilizable"""
    # Encabezados
    for cell in worksheet[1]:
        cell.style = "hdr"
    # Números (aplicar solo a columnas numéricas detectadas por nombre)
    aplicar_formato_numeros(worksheet)
    # Autoajustar anchos
    ajustar_anchos_columna(worksheet)

def aplicar_formato_hoja_vencimiento(worksheet):
    """Aplica formato especial a la hoja VENCIMIENTO (reutilizando estilos)"""
    aplicar_formato_hoja(worksheet)
    # Resaltar filas de totales
    # Columnas esperadas: ... 5=MONEDA, 6=CLIENTE (según tu salida)
    for row_num in range(2, worksheet.max_row + 1):
        moneda_cell = worksheet.cell(row=row_num, column=5)  # MONEDA
        cliente_cell = worksheet.cell(row=row_num, column=6) # CLIENTE
        val_m = (moneda_cell.value or "").strip()
        val_c = (cliente_cell.value or "").strip()
        if val_m in ['Moneda Local', 'Dólar', 'Euro']:
            for col_num in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_num, column=col_num).style = "total_row"
        elif val_c == 'Totales':
            for col_num in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_num, column=col_num).style = "grand_total_row"

def aplicar_formato_numeros(worksheet):
    """Aplica formato de números con separador de miles (NamedStyle 'num') a las columnas relevantes"""
    columnas_numericas = {
        'SALDO_TOTAL', 'SALDO TOTAL', 'SALDO_NO_VENCIDO', 'Saldo No vencido',
        'VENCIDO_30', 'VENCIDO_60', 'VENCIDO_90', 'VENCIDO_180', 'VENCIDO_360', 'VENCIDO_360_PLUS',
        'Vencido 30', 'Vencido 60', 'Vencido 90', 'Vencido 180', 'Vencido 360', 'Vencido + 360',
        'DEUDA_INCOBRABLE', 'DEUDA INCOBRABLE'
    }
    header_map = {}  # col_idx -> is_numeric
    for col_num in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=1, column=col_num)
        header_value = str(header_cell.value) if header_cell.value else ""
        header_map[col_num] = header_value in columnas_numericas
    # Asignar estilo solo a columnas numéricas
    for col_num, is_num in header_map.items():
        if not is_num:
            continue
        for row_num in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            # Evitar tocar celdas vacías
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                continue
            # Si es número o puede serlo, asignar estilo 'num'
            if isinstance(cell.value, (int, float)):
                cell.style = "num"
            else:
                # openpyxl guarda strings cuando vienen de pandas; intenta convertir
                try:
                    float(str(cell.value).replace(",", "").replace(" ", ""))
                    cell.style = "num"
                except:
                    pass  # no forzar

def ajustar_anchos_columna(worksheet):
    """Ajusta automáticamente los anchos de las columnas"""
    for column in worksheet.columns:
        max_length = 0
        first_cell = next(iter(column))
        column_letter = get_column_letter(first_cell.column)
        for cell in column:
            try:
                txt = "" if cell.value is None else str(cell.value)
                if len(txt) > max_length:
                    max_length = len(txt)
            except:
                pass
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# =======================
# CLI (main)
# =======================

def main():
    """Función principal para CLI"""
    import sys
    from pathlib import Path

    if len(sys.argv) < 5:
        print("=" * 60)
        print("MODELO DE DEUDA - PROCESADOR")
        print("=" * 60)
        print("Uso:")
        print("  python modelo_deuda.py <cartera> <anticipos> <TRM_USD> <TRM_EUR> [salida]")
        print("")
        print("Ejemplos:")
        print("  python modelo_deuda.py cartera.csv anticipos.csv 4250.50 4850.75")
        print("  python modelo_deuda.py cartera.xlsx anticipos.xlsx 4250,50 4850,75 resultado.xlsx")
        sys.exit(1)

    cartera_path = sys.argv[1]
    anticipos_path = sys.argv[2]
    
    try:
        trm_usd_str = str(sys.argv[3]).replace('.', '').replace(',', '.')
        trm_eur_str = str(sys.argv[4]).replace('.', '').replace(',', '.')
        trm_usd = float(trm_usd_str)
        trm_eur = float(trm_eur_str)
    except ValueError:
        print("❌ Error: TRM inválidas. Use números como 4250.50 o 4250,50")
        sys.exit(2)

    if len(sys.argv) >= 6:
        output_path = sys.argv[5]
    else:
        base_dir = Path(__file__).resolve().parent
        output_dir = base_dir / 'PROVCA_PROCESADOS'
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = str(output_dir / f'1_Modelo_Deuda_{timestamp}.xlsx')

    try:
        print("🚀 INICIANDO PROCESAMIENTO DEL MODELO DE DEUDA")
        print("=" * 60)
        
        resultado = procesar_modelo_deuda(cartera_path, anticipos_path, trm_usd, trm_eur, output_path)
        
        print("=" * 60)
        print("✅ PROCESAMIENTO COMPLETADO EXITOSAMENTE")
        print(f"📁 Archivo generado: {resultado}")
        print(f"💱 TRM utilizadas - USD: {trm_usd:,.2f} | EUR: {trm_eur:,.2f}")
        
    except Exception as e:
        print("=" * 60)
        print(f"❌ ERROR EN PROCESAMIENTO: {e}")
        sys.exit(3)

if __name__ == "__main__":
    main()
