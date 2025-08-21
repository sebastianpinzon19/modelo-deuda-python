#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador Unificado de Cartera (solo pandas) - con menú

Entradas (en una misma carpeta):
- Balance (ej: contiene "balance")
- Situación (ej: contiene "situacion")
- Focus (ej: contiene "focus")
- Dotación/Provisión del mes (ej: contiene "prov" o "dotacion")
- Acumulado (ej: contiene "acumulado")

Salidas:
- Un Excel consolidado con 7 hojas:
  Balance_Normalizado, Balance_Sumas_Cuentas, Situacion_Total_01010,
  Focus_Vencimientos, Dotacion_Mes, Acumulado, Resumen_Final

Notas:
- Este script prioriza robustez: detección automática de columnas comunes
  y normalización de formatos numéricos (latam).
- Donde el documento Word es ambiguo, se dejan supuestos documentados
  y constantes ajustables en CONFIG.
"""

import os
import re
import sys
import pandas as pd
from datetime import datetime
from trm_config import load_trm, save_trm, parse_trm_value, format_trm_display

# =========================
# CONFIGURACIÓN AJUSTABLE
# =========================

CONFIG = {
    # Palabras clave para encontrar archivos dentro de la carpeta
    "files_glob": {
        "balance": ["balance"],
        "situacion": ["situacion"],
        "focus": ["focus"],
        "dotacion": ["provca", "provision", "dotacion", "prov_"],
        "acumulado": ["acumulado"]
    },
    # Mapeo/heurísticas de columnas típicas:
    "columns": {
        "balance": {
            "division_like": ["DIV", "DIVISION"],
            "cuenta_like": ["CUENTA", "COD"],
            "saldo_like": ["SALDO", "VALOR", "DEBITO", "CRÉDITO", "CREDITO"],
            # Preferir esta columna si existe
            "saldo_prefer_tokens_all": [
                ["saldo", "aaf", "vari"],  # "Saldo AAF variación"
                ["saldo", "aaf"]
            ]
        },
        "situacion": {
            "saldo_mes_like": ["SALDO", "MES"],  # columna que contenga ambas palabras
            "total_row_match": ["TOTAL", "01010"] # fila que contenga ambas
        },
        "focus": {
            # buckets que intentaremos detectar en cualquier orden
            "aging_like": {
                "no_vencido_cols_any": [
                    "por_vencer", "no vencido", "no_vencida", "no_vencidas"
                ],
                "vencido_30_cols_any": ["vencido 30", "30 dias", "30 días", "1-30"],
                "vencido_60_cols_any": ["60 dias", "60 días", "31-60"],
                "vencido_90_cols_any": ["90 dias", "90 días", "61-90"],
                "vencido_120_cols_any": ["120 dias", "120 días", "91-120"],
                "vencido_mas_90_cols_any": ["+90", "mas de 90", "más de 90", ">=90"],
                "vencido_mas_120_cols_any": ["+120", "mas de 120", "más de 120", ">=120"]
            }
        },
        "dotacion": {
            "interco_like": ["interco resto", "interco"],
            "acumuladas_inicial_like": ["dotaciones acumuladas", "inicial"],
            "provision_mes_like": ["provision del mes", "provision mes", "provisión del mes"]
        },
        "acumulado": {
            "fila_objetivo": 54,       # fila 54 (1-based)
            "tomar_columnas_n": 5      # columnas B..F ~ 5 columnas
        }
    },
    # Reglas del Word (supuestos):
    "word_rules": {
        # 4. Cobro de mes - Vencida = (Deuda bruta NO Grupo (Inicial) vencidas - total vencido >=60) / 1000
        # Supuesto: "Deuda bruta ... vencidas" = suma de TODOS los buckets vencidos (>=1 día)
        "cobro_vencida_divisor": 1000.0,

        # 5. Cobro del mes - Total Deuda = (COBROS SITUACION (SALDO MES)) / -1000
        # Supuesto: usamos TOTAL_01010_SALDO_MES de Situación como "COBROS SITUACION (SALDO MES)"
        "cobro_total_divisor": -1000.0,

        # 6. Cobros del mes - No Vencida = Cobro Total - Cobro Vencida
        "cobro_no_vencida_is_residual": True,

        # 7. +/- Vencidos en el mes – vencido = VENCIDO MES 30 días (signo positivo)
        # Supuesto: usamos el bucket "vencido 30" de Focus como ese valor.
        "usar_vencido_30_como_ajuste_mes": True,

        # 10. + Facturación del mes – vencida = 0
        "facturacion_mes_vencida": 0.0,

        # 11. + Facturación del mes – no vencida = (Q22 - H22)  (ambigüo)
        # Supuesto: Para no romper, por defecto 0; ajusta si tienes referencia clara.
        "facturacion_mes_no_vencida": 0.0
    }
}

# =====================
# Utilidades generales
# =====================

def ts():
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")

def asegurar_dir(path):
    os.makedirs(path, exist_ok=True)

def convertir_valor(valor):
    """Convierte strings con formato LATAM a float."""
    try:
        if valor is None or (isinstance(valor, float) and pd.isna(valor)):
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)
        s = str(valor).strip().replace('\u200b','').replace(' ','')
        if s == '' or s.lower() == 'nan':
            return 0.0
        # si tiene ambos separadores, asume latam
        if '.' in s and ',' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s and '.' not in s:
            s = s.replace(',', '.')
        # quita cualquier cosa rara
        s = re.sub(r'[^0-9\.\-]', '', s)
        if s in ('', '-', '.', '-.'):
            return 0.0
        return float(s)
    except:
        return 0.0

def any_in(text, candidates):
    t = str(text).lower()
    return any(c in t for c in candidates)

def find_first_col(df_cols, require_all=None, require_any=None):
    """
    Busca primera columna que cumpla:
      - require_all: lista de palabras que deben estar
      - require_any: lista de palabras de las cuales al menos 1 debe estar
    """
    for c in df_cols:
        col = str(c).strip().lower()
        if require_all and not all(w in col for w in require_all):
            continue
        if require_any and not any(w in col for w in require_any):
            continue
        return c
    return None

def find_cols_any(df_cols, needles_any):
    """Devuelve todas las columnas que contengan alguna de las palabras (needles_any)."""
    found = []
    for c in df_cols:
        col = str(c).lower()
        if any(n in col for n in needles_any):
            found.append(c)
    return found

# =====================
# Utilidades Excel crudas
# =====================

def _col_letter_to_index0(letter):
    """Convierte letra(s) de columna Excel (e.g., 'A', 'Q', 'AA') a índice 0-based."""
    s = str(letter).strip().upper()
    value = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'):
            return None
        value = value * 26 + (ord(ch) - ord('A') + 1)
    return value - 1 if value > 0 else None

def _col_index0_to_letter(index0):
    """Convierte índice 0-based a letra(s) de columna Excel (e.g., 0->A, 1->B, 26->AA)."""
    n = int(index0) + 1
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters

def extraer_facturacion_no_vencida_focus_es(ruta_excel, fila=22, col_q="Q", col_h="H"):
    """
    Intenta leer de la hoja 'España' (o similar) el valor Q22 - H22.
    Si falla, devuelve 0.0. Usa lectura sin encabezados para preservar coordenadas.
    """
    try:
        xls = pd.ExcelFile(ruta_excel)
        hoja_esp = None
        for hoja in xls.sheet_names:
            h = str(hoja).lower()
            if "españa" in h or "espana" in h or "espa" in h:
                hoja_esp = hoja
                break
        if hoja_esp is None:
            hoja_esp = xls.sheet_names[0]

        df = pd.read_excel(ruta_excel, sheet_name=hoja_esp, header=None, dtype=str)
        r = fila - 1
        c_q = _col_letter_to_index0(col_q)
        c_h = _col_letter_to_index0(col_h)
        if r < 0 or r >= len(df.index):
            return 0.0
        vq = convertir_valor(df.iat[r, c_q]) if c_q is not None and c_q < df.shape[1] else 0.0
        vh = convertir_valor(df.iat[r, c_h]) if c_h is not None and c_h < df.shape[1] else 0.0
        return float(vq - vh)
    except Exception:
        return 0.0

# =====================
# Procesadores
# =====================

def leer_excel_multi_hoja(ruta_excel, dtype=str):
    xls = pd.ExcelFile(ruta_excel)
    frames = []
    for hoja in xls.sheet_names:
        try:
            df = pd.read_excel(ruta_excel, sheet_name=hoja, dtype=dtype)
            if not df.empty:
                # normalizar encabezados
                df.columns = [str(c).strip().upper() for c in df.columns]
                df["__HOJA__"] = hoja
                frames.append(df)
        except Exception:
            continue
    if not frames:
        raise ValueError(f"No se pudieron leer hojas válidas de: {os.path.basename(ruta_excel)}")
    return pd.concat(frames, ignore_index=True)

def procesar_balance(ruta_excel):
    conf = CONFIG["columns"]["balance"]
    df = leer_excel_multi_hoja(ruta_excel, dtype=str)
    cols = df.columns.tolist()

    col_div = find_first_col(cols, require_any=[w.lower() for w in conf["division_like"]]) or cols[0]
    col_cuenta = find_first_col(cols, require_any=[w.lower() for w in conf["cuenta_like"]]) or cols[1]
    # Preferir "Saldo AAF variación" si está disponible
    col_saldo = None
    for tokens in conf.get("saldo_prefer_tokens_all", []):
        cand = find_first_col(cols, require_all=tokens)
        if cand:
            col_saldo = cand
            break
    if not col_saldo:
        col_saldo = find_first_col(cols, require_any=[w.lower() for w in conf["saldo_like"]]) or cols[-1]

    out = df[[col_div, col_cuenta, col_saldo]].copy()
    out.columns = ["DIVISION", "CUENTA", "SALDO"]
    out["SALDO"] = out["SALDO"].apply(convertir_valor)
    out = out.groupby(["DIVISION", "CUENTA"], as_index=False)["SALDO"].sum()
    return out

def sumar_cuentas_balance(df_balance):
    """
    Del Word:
      Total cuenta objeto 43001
      0080.43002.20 / .21 / .15 / .28 / .31 / .63
      Total cuenta objeto 43008
      Total cuenta objeto 43042

    Heurística:
    - Si CUENTA contiene exactamente "43001" -> sumar
    - Si CUENTA contiene "43002.20", etc. -> sumar
    - Si CUENTA contiene "43008", "43042" -> sumar
    - También se prueba con prefijo "0080." tal como aparece.
    """
    patrones = {
        # Usamos límites de palabra para evitar grupos de captura y warnings
        "Total_43001": [r"\b43001\b"],
        "Total_43008": [r"\b43008\b"],
        "Total_43042": [r"\b43042\b"],
        "Total_43002_detalle": [
            r"43002\.20", r"43002\.21", r"43002\.15",
            r"43002\.28", r"43002\.31", r"43002\.63",
            r"0080\.43002\.20", r"0080\.43002\.21", r"0080\.43002\.15",
            r"0080\.43002\.28", r"0080\.43002\.31", r"0080\.43002\.63"
        ]
    }

    resumen = []
    for etiqueta, pats in patrones.items():
        mask = df_balance["CUENTA"].astype(str).str.contains("|".join(pats), regex=True, case=False, na=False)
        total = df_balance.loc[mask, "SALDO"].sum()
        resumen.append({"Concepto": etiqueta, "Valor": total})

    return pd.DataFrame(resumen)

def procesar_situacion(ruta_excel):
    conf = CONFIG["columns"]["situacion"]
    df = leer_excel_multi_hoja(ruta_excel, dtype=str)

    # localizar columna "Saldos Mes"
    col_saldo_mes = find_first_col(
        df.columns,
        require_all=[w.lower() for w in conf["saldo_mes_like"]]
    )
    if not col_saldo_mes:
        raise ValueError("Situación: no se encontró columna tipo 'Saldos Mes'.")

    # localizar fila TOTAL 01010
    df_norm = df.astype(str).apply(lambda col: col.str.strip().str.upper())
    mask = df_norm.apply(
        lambda fila: all(k in " ".join(fila.values) for k in conf["total_row_match"]),
        axis=1
    )
    if not mask.any():
        raise ValueError("Situación: no se encontró fila 'TOTAL 01010'.")

    valor = convertir_valor(df.loc[mask, col_saldo_mes].iloc[0])
    return pd.DataFrame({"Concepto": ["TOTAL_01010_SALDO_MES"], "Valor": [valor]})

def _sum_cols_any(df, keys_any):
    """Suma todas las columnas cuyos nombres contengan alguna palabra de keys_any."""
    if not keys_any:
        return 0.0
    cols = find_cols_any(df.columns, [k.lower() for k in keys_any])
    if not cols:
        return 0.0
    vals = df[cols].applymap(convertir_valor).sum().sum()
    return float(vals)

def procesar_focus(ruta_excel):
    """
    Intenta detectar buckets de vencimientos y sumar:
    - No_Vencido
    - Vencido_30
    - Vencido_60
    - Vencido_90
    - Vencido_120
    - Vencido_+90
    - Vencido_+120

    Luego calcula totales:
    - Total_Vencido = suma de todos los buckets vencidos
    - Total_No_Vencido = No_Vencido
    - Total_Deuda = Total_Vencido + Total_No_Vencido
    """
    conf = CONFIG["columns"]["focus"]["aging_like"]
    # Preferir la hoja que contenga "España" (o variantes) del formato España
    xls = pd.ExcelFile(ruta_excel)
    hoja_esp = None
    for hoja in xls.sheet_names:
        h = str(hoja).lower()
        if "españa" in h or "espana" in h or "espa" in h:
            hoja_esp = hoja
            break
    if hoja_esp is not None:
        df = pd.read_excel(ruta_excel, sheet_name=hoja_esp, dtype=str)
        df.columns = [str(c).strip().upper() for c in df.columns]
    else:
        df = leer_excel_multi_hoja(ruta_excel, dtype=str)

    # Asegura numérico para todo lo que sea posible
    df_num = df.copy()
    for c in df_num.columns:
        df_num[c] = df_num[c].apply(convertir_valor)

    # Sumas por buckets (heurístico por nombre de columna)
    no_vencido = _sum_cols_any(df, conf["no_vencido_cols_any"])
    v30 = _sum_cols_any(df, conf["vencido_30_cols_any"])
    v60 = _sum_cols_any(df, conf["vencido_60_cols_any"])
    v90 = _sum_cols_any(df, conf["vencido_90_cols_any"])
    v120 = _sum_cols_any(df, conf["vencido_120_cols_any"])
    vmas90 = _sum_cols_any(df, conf["vencido_mas_90_cols_any"])
    vmas120 = _sum_cols_any(df, conf["vencido_mas_120_cols_any"])

    # Evitar doble conteo: si hay +90 y 90 separados, tomamos el mayor como representativo
    # (ajusta según tus archivos reales)
    total_vencido_componentes = [v30, v60, v90, v120]
    if vmas120 > 0:
        total_vencido_componentes.append(vmas120)
    elif vmas90 > 0:
        total_vencido_componentes.append(vmas90)

    total_vencido = sum(total_vencido_componentes)
    total_no_vencido = no_vencido
    total_deuda = total_vencido + total_no_vencido

    df_out = pd.DataFrame({
        "Concepto": [
            "No_Vencido", "Vencido_30", "Vencido_60", "Vencido_90", "Vencido_120",
            "Vencido_+90", "Vencido_+120", "Total_Vencido", "Total_No_Vencido", "Total_Deuda"
        ],
        "Valor": [
            no_vencido, v30, v60, v90, v120, vmas90, vmas120,
            total_vencido, total_no_vencido, total_deuda
        ]
    })
    return df_out

def procesar_dotacion(ruta_excel):
    """
    Dotación del mes (Word):
      Dotación = Interco_RESTO - Dotaciones_Acumuladas (Inicial) - Provisión del mes
    """
    df = leer_excel_multi_hoja(ruta_excel, dtype=str)
    cols = df.columns

    interco_cols = find_cols_any(cols, [w.lower() for w in CONFIG["columns"]["dotacion"]["interco_like"]])
    acum_cols = [c for c in cols if all(w in c.lower() for w in CONFIG["columns"]["dotacion"]["acumuladas_inicial_like"])]
    prov_cols = find_cols_any(cols, [w.lower() for w in CONFIG["columns"]["dotacion"]["provision_mes_like"]])

    v_interco = df[interco_cols].applymap(convertir_valor).sum().sum() if interco_cols else 0.0
    v_acum = df[acum_cols].applymap(convertir_valor).sum().sum() if acum_cols else 0.0
    v_prov = df[prov_cols].applymap(convertir_valor).sum().sum() if prov_cols else 0.0
    dotacion_mes = v_interco - v_acum - v_prov

    return pd.DataFrame({
        "Concepto": ["Interco_RESTO", "Dotaciones_Acumuladas_Inicial", "Provision_Mes", "Dotacion_Mes"],
        "Valor": [v_interco, v_acum, v_prov, dotacion_mes]
    })

def procesar_acumulado(ruta_excel):
    """
    Lee la primera hoja del archivo de acumulado SIN encabezados y extrae la fila 54,
    columnas B..F (5 columnas). Devuelve conceptos con letras de columna
    para evitar "UNNAMED".
    """
    conf = CONFIG["columns"]["acumulado"]
    try:
        df = pd.read_excel(ruta_excel, sheet_name=0, header=None, dtype=str)
    except Exception:
        df = pd.read_excel(ruta_excel, header=None, dtype=str)

    fila_idx = conf["fila_objetivo"] - 1
    if fila_idx >= len(df.index):
        fila_idx = len(df.index) - 1 if len(df.index) > 0 else 0

    inicio = 1  # Columna B (0=A, 1=B)
    fin_exclusivo = inicio + max(1, int(conf["tomar_columnas_n"]))

    pares = []
    for j in range(inicio, min(fin_exclusivo, df.shape[1])):
        letra = _col_index0_to_letter(j)
        concepto = f"Fila{conf['fila_objetivo']}_Col_{letra}"
        valor = convertir_valor(df.iat[fila_idx, j])
        pares.append((concepto, valor))

    return pd.DataFrame(pares, columns=["Concepto", "Valor"])

# ================
# Resumen Final
# ================

def construir_resumen(df_situacion, df_focus, df_dotacion, fact_no_vencida_override=None):
    """
    Aplica reglas del Word para construir:
      - Cobro_mes_Vencida
      - Cobro_mes_Total
      - Cobro_mes_No_Vencida
      - +/- Vencidos (vencido / no vencido / total)
      - + Facturación (vencida / no vencida)
    """
    rules = CONFIG["word_rules"]

    # Situación: usamos TOTAL_01010_SALDO_MES como "COBROS SITUACION (SALDO MES)"
    cobros_situacion = float(
        df_situacion.set_index("Concepto").loc["TOTAL_01010_SALDO_MES", "Valor"]
    ) if "TOTAL_01010_SALDO_MES" in df_situacion["Concepto"].values else 0.0

    # Focus: sacamos los totales
    map_focus = df_focus.set_index("Concepto")["Valor"].to_dict()
    total_vencido = float(map_focus.get("Total_Vencido", 0.0))
    total_no_vencido = float(map_focus.get("Total_No_Vencido", 0.0))
    total_deuda = float(map_focus.get("Total_Deuda", total_vencido + total_no_vencido))
    vencido_30 = float(map_focus.get("Vencido_30", 0.0))

    # 4. Cobro mes - Vencida
    deuda_bruta_no_grupo_inicial_vencidas = total_vencido  # sup: todas las vencidas
    total_vencido_60omas = float(map_focus.get("Vencido_60", 0.0)) \
                           + float(map_focus.get("Vencido_90", 0.0)) \
                           + float(map_focus.get("Vencido_120", 0.0)) \
                           + max(float(map_focus.get("Vencido_+120", 0.0)), float(map_focus.get("Vencido_+90", 0.0)))
    cobro_mes_vencida = (deuda_bruta_no_grupo_inicial_vencidas - total_vencido_60omas) / rules["cobro_vencida_divisor"]

    # 5. Cobro mes - Total Deuda
    cobro_mes_total = cobros_situacion / rules["cobro_total_divisor"]

    # 6. Cobros del mes - No Vencida = Total - Vencida
    cobro_mes_no_vencida = cobro_mes_total - cobro_mes_vencida if rules["cobro_no_vencida_is_residual"] else 0.0

    # 7. +/- Vencidos en el mes – vencido = VENCIDO MES 30 días
    ajuste_vencidos_mes_vencido = vencido_30 if rules["usar_vencido_30_como_ajuste_mes"] else 0.0

    # 8. +/- No vencido = opuesto del vencido 30 días (según especificación D17)
    ajuste_vencidos_mes_no_vencido = -ajuste_vencidos_mes_vencido

    # 9. +/- Vencidos en el mes – Total deuda = vencido - no vencido
    ajuste_vencidos_mes_total = ajuste_vencidos_mes_vencido - ajuste_vencidos_mes_no_vencido

    # 10 y 11. Facturación del mes
    # Vencida = 0 (según especificación)
    fact_vencida = rules["facturacion_mes_vencida"]
    # No vencida: si recibimos override desde el archivo Focus (Q22-H22), úsalo
    fact_no_vencida = (
        float(fact_no_vencida_override)
        if fact_no_vencida_override is not None
        else rules["facturacion_mes_no_vencida"]
    )

    # Dotación ya viene calculada en df_dotacion
    dotacion_mes = float(df_dotacion.set_index("Concepto").get("Valor").get("Dotacion_Mes", 0.0)) \
        if "Dotacion_Mes" in df_dotacion["Concepto"].values else 0.0

    df_resumen = pd.DataFrame({
        "Concepto": [
            "Cobro_Mes_Total", "Cobro_Mes_Vencida", "Cobro_Mes_No_Vencida",
            "+/- Vencidos_Mes_Vencido", "+/- Vencidos_Mes_No_Vencido", "+/- Vencidos_Mes_Total",
            "Facturacion_Mes_Vencida", "Facturacion_Mes_No_Vencida",
            "Total_Vencido", "Total_No_Vencido", "Total_Deuda",
            "Dotacion_Mes"
        ],
        "Valor": [
            cobro_mes_total, cobro_mes_vencida, cobro_mes_no_vencida,
            ajuste_vencidos_mes_vencido, ajuste_vencidos_mes_no_vencido, ajuste_vencidos_mes_total,
            fact_vencida, fact_no_vencida,
            total_vencido, total_no_vencido, total_deuda,
            dotacion_mes
        ]
    })
    return df_resumen


def construir_resumen_matriz(df_situacion, df_focus, df_dotacion, fact_no_vencida_override=None):
    """
    Devuelve una tabla en formato matriz con columnas Vencida/No_Vencida/Total
    que agrupa los rubros principales del mes, alineada al formato del documento.
    """
    df_r = construir_resumen(df_situacion, df_focus, df_dotacion, fact_no_vencida_override)
    m = df_r.set_index("Concepto")["Valor"].to_dict()

    rows = []
    def add_row(nombre, v_vencida, v_no_vencida):
        rows.append({
            "Concepto": nombre,
            "Vencida": float(v_vencida),
            "No_Vencida": float(v_no_vencida),
            "Total": float(v_vencida) + float(v_no_vencida)
        })

    # Líneas principales
    add_row("Cobros", m.get("Cobro_Mes_Vencida", 0.0), m.get("Cobro_Mes_No_Vencida", 0.0))
    add_row("+ Facturación", m.get("Facturacion_Mes_Vencida", 0.0), m.get("Facturacion_Mes_No_Vencida", 0.0))
    add_row("+/- Vencidos", m.get("+/- Vencidos_Mes_Vencido", 0.0), m.get("+/- Vencidos_Mes_No_Vencido", 0.0))

    # Totales de cartera (informativos)
    rows.append({
        "Concepto": "Total Cartera",
        "Vencida": float(m.get("Total_Vencido", 0.0)),
        "No_Vencida": float(m.get("Total_No_Vencido", 0.0)),
        "Total": float(m.get("Total_Deuda", 0.0))
    })

    # Dotación (se muestra en la columna Total)
    rows.append({
        "Concepto": "Dotacion del Mes",
        "Vencida": 0.0,
        "No_Vencida": 0.0,
        "Total": float(m.get("Dotacion_Mes", 0.0))
    })

    return pd.DataFrame(rows, columns=["Concepto", "Vencida", "No_Vencida", "Total"])

# =====================
# Búsqueda de archivos
# =====================

def pick_file(carpeta, patterns):
    files = os.listdir(carpeta)
    for f in files:
        name = f.lower()
        if any(p in name for p in patterns):
            if f.lower().endswith((".xlsx", ".xls", ".csv")):
                return os.path.join(carpeta, f)
    return None

def descubrir_archivos(carpeta):
    globs = CONFIG["files_glob"]
    balance = pick_file(carpeta, globs["balance"])
    situacion = pick_file(carpeta, globs["situacion"])
    focus = pick_file(carpeta, globs["focus"])
    dotacion = pick_file(carpeta, globs["dotacion"])
    acumulado = pick_file(carpeta, globs["acumulado"])
    return balance, situacion, focus, dotacion, acumulado

# =====================
# Guardado Excel final
# =====================

def guardar_excel_salida(carpeta, **dfs):
    # Usar la carpeta PROVCA_PROCESADOS en lugar de crear una subcarpeta
    out_dir = r"C:\wamp64\www\modelo-deuda-python\cartera_v2.0.0\PROVCA_PROCESADOS"
    asegurar_dir(out_dir)
    out_file = os.path.join(out_dir, f"procesamiento_unificado_{ts()}.xlsx")
    # Auto-ajuste de columnas y formato numérico usando openpyxl
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        for nombre_hoja, df in dfs.items():
            sheet_name = nombre_hoja[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Congelar encabezados
            ws.freeze_panes = "A2"

            # Auto ancho de columnas y formato numérico para evitar notación científica
            for idx_col, col in enumerate(df.columns, start=1):
                # Calcular ancho
                series = df[col]
                max_len = max(
                    len(str(col)),
                    max((len(str(x)) for x in series.fillna("")), default=0)
                )
                ws.column_dimensions[get_column_letter(idx_col)].width = max(10, min(max_len + 2, 60))

                # Formato numérico para toda la columna (filas de datos)
                for row in range(2, 2 + len(df)):
                    cell = ws.cell(row=row, column=idx_col)
                    if isinstance(cell.value, (int, float)):
                        # Mostrar con miles y 2 decimales para evitar 1E-09 y similares
                        cell.number_format = '#,##0.00'
    return out_file

# ================
# Menú principal
# ================

def ejecutar_proceso(carpeta):
    balance_file, situacion_file, focus_file, dotacion_file, acumulado_file = descubrir_archivos(carpeta)
    faltantes = [n for n, p in [
        ("Balance", balance_file),
        ("Situacion", situacion_file),
        ("Focus", focus_file),
        ("Dotacion/Provisión", dotacion_file),
        ("Acumulado", acumulado_file),
    ] if not p]

    if faltantes:
        raise FileNotFoundError("No se encontraron estos archivos en la carpeta: " + ", ".join(faltantes))

    # Aviso de TRM vigente (persistida)
    cfg = load_trm()
    print(f"TRM actual guardada -> USD: {cfg.get('usd')} | EUR: {cfg.get('eur')} (fecha: {cfg.get('updated_at')})")

    # --- Procesos ---
    df_balance = procesar_balance(balance_file)
    df_balance_sumas = sumar_cuentas_balance(df_balance)
    df_situacion = procesar_situacion(situacion_file)
    df_focus = procesar_focus(focus_file)
    df_dotacion = procesar_dotacion(dotacion_file)
    df_acumulado = procesar_acumulado(acumulado_file)

    # Cálculo específico: Facturación del mes – No vencida = Q22 - H22 (hoja España)
    fact_no_vencida_es = extraer_facturacion_no_vencida_focus_es(focus_file, fila=22, col_q="Q", col_h="H")
    df_resumen = construir_resumen(df_situacion, df_focus, df_dotacion, fact_no_vencida_override=fact_no_vencida_es)

    # Guardar
    out_file = guardar_excel_salida(carpeta,
        Balance_Normalizado=df_balance,
        Balance_Sumas_Cuentas=df_balance_sumas,
        Situacion_Total_01010=df_situacion,
        Focus_Vencimientos=df_focus,
        Dotacion_Mes=df_dotacion,
        Acumulado=df_acumulado,
        Resumen_Final=df_resumen,
        Resumen_Matriz=construir_resumen_matriz(df_situacion, df_focus, df_dotacion, fact_no_vencida_override=fact_no_vencida_es)
    )
    return out_file


def menu():
    while True:
        print("\n=== MENÚ PROCESADOR UNIFICADO ===")
        print("1. Procesar carpeta de datos")
        print("2. Ver/Configurar TRM (Dólar/Euro)")
        print("3. Salir")
        opcion = input("Seleccione una opción: ").strip()

        if opcion == "1":
            carpeta = input("Ruta de la carpeta con los 5 archivos: ").strip('"')
            if not os.path.isdir(carpeta):
                print("⚠ La ruta no es una carpeta válida.")
                continue
            try:
                print("🔄 Procesando...")
                out = ejecutar_proceso(carpeta)
                print(f"✅ Listo. Archivo generado en:\n{out}")
            except Exception as e:
                print(f"❌ Error: {e}")
        elif opcion == "2":
            cfg = load_trm()
            print(f"TRM guardadas → USD: {format_trm_display(cfg.get('usd'))} | EUR: {format_trm_display(cfg.get('eur'))} (actualizado: {cfg.get('updated_at')})")
            print("Deje vacío para mantener el valor actual.")
            usd_txt = input("Nueva TRM Dólar: ").strip()
            eur_txt = input("Nueva TRM Euro: ").strip()
            try:
                new_usd = parse_trm_value(usd_txt, cfg.get('usd'))
                new_eur = parse_trm_value(eur_txt, cfg.get('eur'))
                save_trm(new_usd, new_eur)
                print(f"✅ TRM guardadas. USD: {format_trm_display(new_usd)} | EUR: {format_trm_display(new_eur)}")
            except Exception as e:
                print(f"❌ No se pudieron guardar las TRM: {e}")
        elif opcion == "3":
            print("👋 Saliendo...")
            sys.exit(0)
        else:
            print("⚠ Opción no válida, intente de nuevo.")

if __name__ == "__main__":
    if len(sys.argv) == 2 and os.path.isdir(sys.argv[1]):
        try:
            ruta = ejecutar_proceso(sys.argv[1])
            print(f"✅ Listo. Archivo generado en:\n{ruta}")
        except Exception as e:
            print(f"❌ Error: {e}")
            sys.exit(1)
    else:
        menu()
