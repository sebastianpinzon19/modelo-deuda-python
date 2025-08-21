import json
import os
from glob import glob
import openpyxl
import json

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
OUT_DIR = os.path.join(ROOT, 'PROVCA_PROCESADOS')

def latest_model_file():
    files = sorted(glob(os.path.join(OUT_DIR, '1_Modelo_Deuda_*.xlsx')), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

def analyze_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    out = { 'headers': headers }

    def find(col):
        try:
            return headers.index(col) + 1
        except ValueError:
            return -1

    # check numeric formats and thousand separators presence
    first_num_col = None
    for cand in ['SALDO', 'SALDO TOTAL', 'SALDO NO VENCIDO', 'VENCIDO 30']:
        idx = find(cand)
        if idx > 0:
            first_num_col = idx
            break
    if first_num_col:
        # find a non-empty value row to inspect format
        fmt = None
        sample_vals = []
        for r in range(2, min(100, ws.max_row + 1)):
            v = ws.cell(row=r, column=first_num_col).value
            if isinstance(v, (int, float)):
                fmt = ws.cell(row=r, column=first_num_col).number_format
                sample_vals.append(v)
                break
        out['num_format_sample'] = fmt
        out['sample_value'] = sample_vals[0] if sample_vals else None

    # empty columns detection (first 200 rows)
    empty_cols = []
    for idx, h in enumerate(headers, start=1):
        if not h:
            empty_cols.append(h)
            continue
        nonempty = False
        for r in range(2, min(ws.max_row, 200) + 1):
            v = ws.cell(row=r, column=idx).value
            if v not in (None, ''):
                nonempty = True
                break
        if not nonempty:
            empty_cols.append(h)
    out['empty_columns'] = empty_cols
    return out

def main():
    modelo = latest_model_file()
    if not modelo:
        print(json.dumps({'error': 'NO_MODEL_FOUND', 'path': OUT_DIR}))
        return
    wb = openpyxl.load_workbook(modelo, data_only=True)
    res = {'file': modelo, 'sheets': wb.sheetnames}
    for sheet in ['PESOS', 'DIVISAS', 'VENCIMIENTO']:
        if sheet in wb.sheetnames:
            res[sheet] = analyze_sheet(wb, sheet)
    # extra: TRM column presence in VENCIMIENTO
    if 'VENCIMIENTO' in wb.sheetnames:
        ws = wb['VENCIMIENTO']
        headers = [c.value for c in ws[1]]
        res['VENCIMIENTO']['has_TRM'] = 'TRM' in headers
        # last labels in CLIENTE column to verify totals
        try:
            col_cli = headers.index('CLIENTE') + 1
            labels = []
            for r in range(max(2, ws.max_row - 5), ws.max_row + 1):
                labels.append(ws.cell(row=r, column=col_cli).value)
            res['VENCIMIENTO']['last_labels'] = labels
        except ValueError:
            res['VENCIMIENTO']['last_labels'] = []
    # Extra: validar sumatoria de buckets en VENCIMIENTO
    if 'VENCIMIENTO' in wb.sheetnames:
        ws = wb['VENCIMIENTO']
        headers = [c.value for c in ws[1]]
        def idx(h):
            try: return headers.index(h) + 1
            except ValueError: return -1
        col_total = idx('SALDO TOTAL')
        bucket_cols = [idx(h) for h in ['SALDO NO VENCIDO','VENCIDO 30','VENCIDO 60','VENCIDO 90','VENCIDO 180','VENCIDO 360','VENCIDO + 360'] if idx(h)>0]
        diffs = []
        for r in range(2, min(ws.max_row, 500)+1):
            try:
                total = ws.cell(row=r, column=col_total).value or 0
                suma = 0
                for c in bucket_cols:
                    v = ws.cell(row=r, column=c).value or 0
                    if isinstance(v, str):
                        continue
                    suma += v
                if isinstance(total, (int,float)) and (total - suma) != 0:
                    diffs.append({'row': r, 'diff': total - suma})
            except Exception:
                continue
        res['VENCIMIENTO']['bucket_diffs_count'] = len(diffs)
    print(json.dumps(res, ensure_ascii=False))

if __name__ == '__main__':
    main()


